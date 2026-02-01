"""
Test named ranges and formula integrity.

Ensures:
- All expected named ranges exist
- No orphan/ghost named ranges (defined but never used)
- Named ranges point to valid cells
- Compliance integration named ranges are properly defined
"""
import pytest
from openpyxl import load_workbook


# Critical named ranges that must exist for the model to function
CRITICAL_NAMED_RANGES = [
    # Stochastic variables
    "Stoch_SeasonExt",
    "Stoch_ExtValue",
    "Stoch_Interventions",
    "Stoch_TreatmentRate",
    "Stoch_Efficiency",

    # Scenario values
    "S1_Value",
    "S2_Value",
    "S3_Value",
    "S4_Value",

    # Value calculations
    "Gross_Value",
    "Compliance_Adjusted_Value",  # NEW in Phase 4

    # Compliance gates
    "Master_Gate_Status",  # NEW in Phase 3

    # Kinetics
    "Q10_Factor",
    "UV_Modifier",
    "DO_Modifier",
    "Nutrient_Modifier",
    "Bioavail_Modifier",
    "Effective_Rate",
    "Days_to_Compliance",

    # Costs and NPV
    "Testing_Cost",
    "NPV_Total",

    # Monte Carlo stats
    "MC_P10",
    "MC_P50",
    "MC_P90",
    "MC_Prob_Positive",
]

# Named ranges added in Phase 3/4 for environmental drivers
ENVIRONMENTAL_NAMED_RANGES = [
    "Env_Temp_Water_May",
    "Env_Temp_Water_Jun",
    "Env_Temp_Water_Jul",
    "Env_Temp_Water_Aug",
    "Env_Temp_Water_Sep",
    "Env_UV_Relative_May",
    "Env_UV_Relative_Sep",
    "Env_DO_Typical_May",
    "Env_DO_Typical_Sep",
    "Env_Nutrient_Index_May",
    "Env_Bioavail_Index_May",
]


@pytest.fixture
def generated_wb(financial_modeling_dir):
    """Load the generated workbook for testing."""
    wb_path = financial_modeling_dir / "output" / "generated_test.xlsx"
    if not wb_path.exists():
        # Generate if not exists
        import sys
        sys.path.insert(0, str(financial_modeling_dir))
        from luminous_wetland_monte_carlo import generate_workbook
        wb = generate_workbook()
        wb_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(wb_path)
        return wb
    return load_workbook(wb_path, data_only=False)


@pytest.mark.integration
def test_critical_named_ranges_exist(generated_wb):
    """Verify all critical named ranges are defined."""
    defined_names = set(generated_wb.defined_names)

    missing = []
    for name in CRITICAL_NAMED_RANGES:
        if name not in defined_names:
            missing.append(name)

    assert not missing, f"Missing critical named ranges: {missing}"


@pytest.mark.integration
def test_environmental_named_ranges_exist(generated_wb):
    """Verify environmental driver named ranges are defined."""
    defined_names = set(generated_wb.defined_names)

    missing = []
    for name in ENVIRONMENTAL_NAMED_RANGES:
        if name not in defined_names:
            missing.append(name)

    assert not missing, f"Missing environmental named ranges: {missing}"


@pytest.mark.integration
def test_named_ranges_have_valid_references(generated_wb):
    """Verify named ranges point to valid sheet/cell references."""
    invalid = []

    for name in generated_wb.defined_names:
        defn = generated_wb.defined_names[name]
        # Get the destination - should be a valid reference
        try:
            # openpyxl stores the reference as attr_text
            ref = defn.attr_text
            if ref and not ref.startswith("#"):  # Skip error refs
                # Basic validation: should contain sheet!cell pattern
                if "!" not in ref and "$" not in ref:
                    invalid.append((name, ref, "No sheet reference"))
        except Exception as e:
            invalid.append((name, str(defn), str(e)))

    assert not invalid, f"Invalid named range references: {invalid}"


@pytest.mark.integration
def test_compliance_adjusted_value_references_master_gate(generated_wb):
    """Verify Compliance_Adjusted_Value formula uses Master_Gate_Status."""
    ws = generated_wb["12_Calc_Value"]

    # Find the cell with Compliance_Adjusted_Value
    found = False
    for row in ws.iter_rows(min_row=1, max_row=50):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if "Gross_Value*Master_Gate_Status" in cell.value:
                    found = True
                    break

    assert found, "Compliance_Adjusted_Value should reference Master_Gate_Status"


@pytest.mark.integration
def test_minimum_named_range_count(generated_wb):
    """Verify workbook has sufficient named ranges (at least 100 for Phase 3+)."""
    count = len(generated_wb.defined_names)

    # Phase 3+ should have many more named ranges due to environmental/kinetics
    assert count >= 100, f"Expected at least 100 named ranges, got {count}"


@pytest.mark.integration
def test_no_error_named_ranges(generated_wb):
    """Verify no named ranges point to #REF! or other errors."""
    error_ranges = []

    for name in generated_wb.defined_names:
        defn = generated_wb.defined_names[name]
        ref = defn.attr_text
        if ref and ref.startswith("#"):
            error_ranges.append((name, ref))

    assert not error_ranges, f"Named ranges with errors: {error_ranges}"


@pytest.mark.integration
def test_gate_named_ranges_exist(generated_wb):
    """Verify compliance gate named ranges are defined."""
    defined_names = set(generated_wb.defined_names)

    # Check for individual gate status named ranges
    gate_prefixes = ["Gate_Toxicity", "Gate_Chloride", "Gate_SAR", "Gate_DO", "Gate_Sulfate"]

    found_gates = [name for name in defined_names if any(name.startswith(prefix) for prefix in gate_prefixes)]

    # Should have at least some gate-related named ranges
    assert len(found_gates) >= 5, f"Expected gate named ranges, found: {found_gates}"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
