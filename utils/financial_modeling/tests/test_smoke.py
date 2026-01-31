"""
Phase 0 Smoke Test - Validates baseline workbook before refactoring.

Run with: pytest tests/test_smoke.py -v
"""
import pytest
from openpyxl import load_workbook


# Expected sheet names for the 22-sheet Phase 3+ workbook (with new environmental/kinetics/compliance sheets)
EXPECTED_SHEETS_CURRENT = [
    "0_Cover",
    "1_TOC",
    "3_Instructions",
    "2_InputMap",
    "4_Assumptions",
    "5_EnvironmentalDrivers",  # NEW in Phase 3
    "6_Scenarios",
    "7_ServiceModels",
    "8_Calc_Timeline",
    "9_Calc_Kinetics",         # NEW in Phase 3
    "10_Calc_Stochastic",
    "11_Calc_Compliance",      # NEW in Phase 3
    "12_Calc_Value",
    "13_Calc_Costs",
    "14_Calc_Sim",
    "15_PL_Monthly",
    "16_PL_Annual",
    "17_CashFlow",
    "18_UnitEconomics",
    "19_Sensitivity",
    "20_Dashboard",
    "21_Checks",
]

# Original baseline sheet names (from fixtures/baseline_workbook.xlsx)
EXPECTED_SHEETS_V1 = [
    "0_Cover",
    "1_TOC",
    "2_Instructions",
    "1.5_InputMap",
    "3_Assumptions",
    "4_Scenarios",
    "5_ServiceModels",
    "6_Calc_Timeline",
    "7_Calc_Stochastic",
    "8_Calc_Value",
    "9_Calc_Costs",
    "10_Calc_Sim",
    "11_PL_Monthly",
    "12_PL_Annual",
    "13_CashFlow",
    "14_UnitEconomics",
    "15_Sensitivity",
    "16_Dashboard",
    "17_Checks",
]


@pytest.mark.integration
def test_baseline_workbook_exists(baseline_workbook_path):
    """Verify baseline workbook exists in fixtures."""
    assert baseline_workbook_path.exists(), (
        f"Baseline workbook not found at {baseline_workbook_path}. "
        "Run Phase 0 setup to create it."
    )


@pytest.mark.integration
def test_baseline_workbook_has_expected_sheet_count(baseline_workbook_path):
    """Verify baseline workbook has exactly 19 sheets (18 + 1.5_InputMap)."""
    wb = load_workbook(baseline_workbook_path, read_only=True)
    sheet_count = len(wb.sheetnames)
    wb.close()

    # 18 sheets including the 1.5_InputMap
    assert sheet_count == 19, (
        f"Expected 19 sheets, got {sheet_count}. Sheets: {wb.sheetnames}"
    )


@pytest.mark.integration
def test_baseline_workbook_has_expected_sheets(baseline_workbook_path):
    """Verify all expected sheet names are present."""
    wb = load_workbook(baseline_workbook_path, read_only=True)
    actual_sheets = set(wb.sheetnames)
    wb.close()

    expected_sheets = set(EXPECTED_SHEETS_V1)
    missing = expected_sheets - actual_sheets
    extra = actual_sheets - expected_sheets

    assert not missing, f"Missing sheets: {missing}"
    assert not extra, f"Unexpected extra sheets: {extra}"


@pytest.mark.integration
def test_baseline_workbook_sheet_order(baseline_workbook_path):
    """Verify sheets are in expected order."""
    wb = load_workbook(baseline_workbook_path, read_only=True)
    actual_sheets = wb.sheetnames
    wb.close()

    assert actual_sheets == EXPECTED_SHEETS_V1, (
        f"Sheet order mismatch.\nExpected: {EXPECTED_SHEETS_V1}\nActual: {actual_sheets}"
    )


@pytest.mark.integration
def test_generated_workbook_has_current_sheets(generated_workbook_path):
    """Verify newly generated workbook has the 22-sheet structure."""
    if not generated_workbook_path.exists():
        pytest.skip("Generated workbook not found - run generator first")

    gen_wb = load_workbook(generated_workbook_path, read_only=True)
    gen_sheets = gen_wb.sheetnames
    gen_wb.close()

    # Check sheet count - now 22 sheets
    assert len(gen_sheets) == 22, (
        f"Expected 22 sheets, got {len(gen_sheets)}. Sheets: {gen_sheets}"
    )

    # Check all expected sheets are present
    expected_set = set(EXPECTED_SHEETS_CURRENT)
    actual_set = set(gen_sheets)
    missing = expected_set - actual_set
    extra = actual_set - expected_set

    assert not missing, f"Missing sheets: {missing}"
    assert not extra, f"Unexpected extra sheets: {extra}"


@pytest.mark.integration
def test_generated_workbook_has_new_sheets(generated_workbook_path):
    """Verify the three new sheets (Environmental, Kinetics, Compliance) exist."""
    if not generated_workbook_path.exists():
        pytest.skip("Generated workbook not found - run generator first")

    gen_wb = load_workbook(generated_workbook_path, read_only=True)
    gen_sheets = set(gen_wb.sheetnames)
    gen_wb.close()

    new_sheets = ["5_EnvironmentalDrivers", "9_Calc_Kinetics", "11_Calc_Compliance"]
    for sheet in new_sheets:
        assert sheet in gen_sheets, f"New sheet '{sheet}' not found in workbook"


@pytest.mark.integration
def test_generated_workbook_has_named_ranges(generated_workbook_path):
    """Verify generated workbook has named ranges defined."""
    if not generated_workbook_path.exists():
        pytest.skip("Generated workbook not found - run generator first")

    wb = load_workbook(generated_workbook_path, read_only=True)
    named_range_count = len(wb.defined_names)
    wb.close()

    # Should have many named ranges (at least 50 based on the script)
    assert named_range_count >= 50, (
        f"Expected at least 50 named ranges, got {named_range_count}"
    )


@pytest.mark.integration
def test_baseline_has_named_ranges(baseline_workbook_path):
    """Verify baseline workbook has named ranges defined."""
    wb = load_workbook(baseline_workbook_path, read_only=True)
    # openpyxl 3.x API: defined_names is a dict-like object
    named_range_count = len(wb.defined_names)
    wb.close()

    # Should have many named ranges (at least 50 based on the script)
    assert named_range_count >= 50, (
        f"Expected at least 50 named ranges, got {named_range_count}"
    )


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
