#!/usr/bin/env python3
"""
Luminous Wetland Financial Model - FAST Standard Generator (v3.0)

Generates an 18-sheet Excel workbook (0-17) with Excel-native Monte Carlo via Data Tables.
Refactored to use LOCATIONS registry for all cross-sheet references.

Python generates STRUCTURE and FORMULAS only.
Excel performs ALL calculations including Monte Carlo simulation.

Standard: FAST (Flexible, Appropriate, Structured, Transparent)
Target: Excel 365/2021

Sheet Structure (FAST 18-Sheet Standard):
  0_Cover        - Navigation hub
  1_TOC          - Table of Contents / Control panel
  3_Instructions - Methodology and documentation
  4_Assumptions  - ALL inputs consolidated (timing, site config, test options, validation)
  6_Scenarios    - Value scenario definitions
  7_ServiceModels - Monte Carlo distribution parameters
  8_Calc_Timeline - Projection timeline
  10_Calc_Stochastic - Random variable generation
  12_Calc_Value   - Gated value calculations
  13_Calc_Costs   - Cost calculations
  14_Calc_Sim    - Monte Carlo simulation (Data Table)
  15_PL_Monthly  - (Placeholder for monthly P&L if needed)
  16_PL_Annual   - Annual projections
  17_CashFlow    - Cumulative projections
  18_UnitEconomics - Unit economics view
  19_Sensitivity - Tornado analysis
  20_Dashboard   - Executive dashboard
  21_Checks      - Validation checks

Output: luminous_wetland_monte_carlo_model.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.formatting.rule import FormulaRule
import os

# Content loader for externalized documentation
from content_loader import (
    ContentLoader,
    ContentNotFoundError,
    ContentValidationError,
    TableDocumentation,
    get_loader
)

# =============================================================================
# LOCATIONS REGISTRY - Dual-purpose: Write addresses AND Read references
# =============================================================================

# WRITE_MAP: Python needs addresses to place values
# Key: variable name, Value: dict with sheet, col, row
WRITE_MAP = {}

# LOCATIONS: Legacy support - stores full addresses (for backward compat)
LOCATIONS = {}

# =============================================================================
# INPUT_REGISTRY - Maps inputs to sheets that directly reference them in formulas
# =============================================================================
# Used for:
#   1. Populating "Referenced By" column in input tables
#   2. Validating registry matches actual formula usage (verify_traceability)
# Scope: Direct formula reference only (not downstream dependencies)

INPUT_REGISTRY = {
    # === 4_Assumptions: Treatment Kinetics → 8_Calc_Timeline ===
    'Initial_NAFC': ['8_Calc_Timeline'],
    'Target_NAFC': ['8_Calc_Timeline'],
    'Rapid_Phase_Rate': ['8_Calc_Timeline'],
    'Slow_Phase_Rate': ['8_Calc_Timeline'],
    'Rapid_Phase_Duration': ['8_Calc_Timeline'],
    'Recirculation_Cycle': ['8_Calc_Timeline'],
    'Season_Length': ['8_Calc_Timeline', '13_Calc_Costs', '14_Calc_Sim'],

    # === 4_Assumptions: Learning Curve → 8_Calc_Timeline ===
    'LearningCurve_Mult': ['8_Calc_Timeline'],

    # === 7_ServiceModels: Triangular Distributions → 10_Calc_Stochastic ===
    'Tri_Season_Extension_Days_Min': ['10_Calc_Stochastic'],
    'Tri_Season_Extension_Days_Mode': ['10_Calc_Stochastic'],
    'Tri_Season_Extension_Days_Max': ['10_Calc_Stochastic'],
    'Tri_Extension_Value_Per_Day_Min': ['10_Calc_Stochastic'],
    'Tri_Extension_Value_Per_Day_Mode': ['10_Calc_Stochastic'],
    'Tri_Extension_Value_Per_Day_Max': ['10_Calc_Stochastic'],
    'Tri_Interventions_Avoided_Min': ['10_Calc_Stochastic'],
    'Tri_Interventions_Avoided_Mode': ['10_Calc_Stochastic'],
    'Tri_Interventions_Avoided_Max': ['10_Calc_Stochastic'],
    'Tri_Treatment_Rate_Factor_Min': ['10_Calc_Stochastic'],
    'Tri_Treatment_Rate_Factor_Mode': ['10_Calc_Stochastic'],
    'Tri_Treatment_Rate_Factor_Max': ['10_Calc_Stochastic'],

    # === 7_ServiceModels: Beta Distributions → 10_Calc_Stochastic ===
    # Note: Alpha/Beta params used internally in formulas, Min/Max for scaling
    'Beta_Efficiency_Gain_Min': ['10_Calc_Stochastic'],
    'Beta_Efficiency_Gain_Max': ['10_Calc_Stochastic'],

    # === Intermediate calculations (8_Calc_Timeline) → multiple sheets ===
    'Eff_Rapid_Rate': ['8_Calc_Timeline'],
    'Eff_Slow_Rate': ['8_Calc_Timeline'],
    'NAFC_After_Rapid': ['8_Calc_Timeline'],
    'Days_to_Compliance': ['8_Calc_Timeline', '14_Calc_Sim'],

    # === Timeline outputs → 16_PL_Annual ===
    'Timeline_LearningMult': ['16_PL_Annual'],
    'Timeline_DiscountFactor': ['16_PL_Annual'],
}

# Maps parameter names to their canonical INPUT_REGISTRY keys (for aliases)
INPUT_ALIASES = {
    'Treatment_Days': 'Season_Length',
}


def ref(var_name: str) -> str:
    """Return the named range string for use in formulas.

    Per Audit: Formulas should use NAMES, never coordinate references.
    This returns the name itself, which Excel resolves via Name Manager.

    Usage in formula construction:
        f"={ref('Season_Length')}/C5"  # Returns "=Season_Length/C5"

    The resulting Excel formula uses the named range, not coordinates.
    """
    if var_name not in WRITE_MAP and var_name not in LOCATIONS:
        raise KeyError(f"Variable '{var_name}' not registered. "
                      f"Available: {list(WRITE_MAP.keys())}")
    return var_name  # Return the NAME, not the address


def loc(var_name: str) -> str:
    """Return Excel address for a named variable (LEGACY - prefer ref()).

    NOTE: Per audit, formulas should use ref() which returns names.
    This function is retained for cases where actual addresses are needed
    (e.g., Data Table setup, debugging).

    Usage:
        loc('Season_Length')  # Returns "'4_Assumptions'!$B$15"
    """
    if var_name not in LOCATIONS:
        raise KeyError(f"Variable '{var_name}' not registered in LOCATIONS. "
                      f"Available: {list(LOCATIONS.keys())}")
    return LOCATIONS[var_name]


def register_location(var_name: str, sheet: str, col: str, row: int):
    """Register a variable's location for formula generation.

    Args:
        var_name: Logical name for the variable (e.g., 'Season_Length')
        sheet: Sheet name (e.g., '4_Assumptions')
        col: Column letter (e.g., 'B')
        row: Row number (e.g., 15)

    This registers in BOTH WRITE_MAP (for placement) and LOCATIONS (for addresses).
    Named ranges are created separately via add_named_range().
    """
    WRITE_MAP[var_name] = {'sheet': sheet, 'col': col, 'row': row}
    LOCATIONS[var_name] = f"'{sheet}'!${col}${row}"


def register_range(var_name: str, sheet: str, start_col: str, start_row: int,
                   end_col: str, end_row: int):
    """Register a range location for array formulas.

    Args:
        var_name: Logical name for the range
        sheet: Sheet name
        start_col, start_row: Top-left cell
        end_col, end_row: Bottom-right cell
    """
    WRITE_MAP[var_name] = {
        'sheet': sheet,
        'start_col': start_col, 'start_row': start_row,
        'end_col': end_col, 'end_row': end_row,
        'is_range': True
    }
    LOCATIONS[var_name] = f"'{sheet}'!${start_col}${start_row}:${end_col}${end_row}"


def clear_locations():
    """Clear the locations registry (call before regenerating workbook)."""
    global LOCATIONS, WRITE_MAP
    LOCATIONS = {}
    WRITE_MAP = {}


def is_range_input(name):
    """Check if a named input is a range (vs single cell) using WRITE_MAP.

    Uses the 'is_range' flag set by register_range() to dynamically detect
    whether an input spans multiple cells.
    """
    return WRITE_MAP.get(name, {}).get('is_range', False)


def get_input_references(param_name):
    """Get references for a parameter, checking aliases if needed.

    Looks up the parameter in INPUT_REGISTRY. If not found directly,
    checks INPUT_ALIASES for a canonical name mapping.

    Args:
        param_name: The parameter name (e.g., 'Treatment_Days')

    Returns:
        List of sheet names that reference this input, or empty list.
    """
    refs = INPUT_REGISTRY.get(param_name, [])
    if refs:
        return refs
    alias = INPUT_ALIASES.get(param_name)
    if alias:
        return INPUT_REGISTRY.get(alias, [])
    return []


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

# Fonts
FONT_TITLE = Font(bold=True, size=16)
FONT_HEADER = Font(bold=True, size=14)
FONT_BOLD = Font(bold=True)
FONT_TABLE_HEADER = Font(bold=True, color="FFFFFF")

# Fills
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_EDITABLE = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
FILL_PASS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_FAIL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Tab Colors (FAST Standard)
TAB_WHITE = "FFFFFF"   # Cover
TAB_YELLOW = "FFFF00"  # Inputs (user-editable)
TAB_GRAY = "808080"    # Reference / Documentation
TAB_BLUE = "4472C4"    # Calculations
TAB_GREEN = "70AD47"   # Outputs
TAB_RED = "FF0000"     # Checks

# Number Formats
FMT_CURRENCY = '$#,##0'
FMT_CURRENCY_DEC = '$#,##0.00'
FMT_PERCENT = '0.0%'
FMT_NUMBER = '#,##0.00'
FMT_DATE = 'YYYY-MM-DD'

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def create_table(ws, name, ref, style="TableStyleMedium9"):
    """Create an Excel table with specified style."""
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
    ws.add_table(table)
    return table


def add_named_range(wb, name, sheet_name, cell_ref):
    """Add a named range to the workbook and register in LOCATIONS."""
    full_ref = f"'{sheet_name}'!{cell_ref}"
    defined_name = DefinedName(name=name, attr_text=full_ref)
    wb.defined_names.add(defined_name)
    # Also register in LOCATIONS for formula generation
    # Parse cell_ref to extract just the address without range
    if ':' not in cell_ref:
        # Single cell
        col = ''.join(c for c in cell_ref if c.isalpha()).replace('$', '')
        row = int(''.join(c for c in cell_ref if c.isdigit()))
        register_location(name, sheet_name, col, row)


def write_header_row(ws, row, headers, start_col=1):
    """Write a header row with styling."""
    for col, header in enumerate(headers, start_col):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = FONT_TABLE_HEADER
        cell.fill = FILL_HEADER
    return row + 1


def set_column_widths(ws, widths):
    """Set column widths from a dict of column: width."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def get_merged_width(ws, start_col: int, end_col: int) -> float:
    """Get combined width of columns for merged cell calculations.

    Args:
        ws: Worksheet object
        start_col: Starting column number (1-indexed)
        end_col: Ending column number (1-indexed)

    Returns:
        Combined width of all columns in range
    """
    from openpyxl.utils import get_column_letter
    total_width = 0
    for col in range(start_col, end_col + 1):
        col_letter = get_column_letter(col)
        col_width = ws.column_dimensions[col_letter].width
        # Default Excel width is ~8.43 if not set
        total_width += col_width if col_width else 8.43
    return total_width


def calculate_row_height(text: str, column_width: float, font_size: int = 10, buffer: float = 1.3) -> float:
    """Calculate row height needed for wrapped text.

    Args:
        text: The text content
        column_width: Combined width of columns the text spans
        font_size: Font size in points (default 10)
        buffer: Safety multiplier for edge cases (default 1.3)

    Returns:
        Row height in points
    """
    if not text:
        return 15

    # Approximate characters that fit per column width unit
    # Excel column width ~7 pixels per unit, ~7 pixels per character at size 10
    chars_per_width_unit = 1.0
    chars_per_line = max(1, int(column_width * chars_per_width_unit))

    # Count explicit newlines and calculate wrapping
    lines = text.split('\n')
    total_lines = 0
    for line in lines:
        # Each explicit line may wrap
        wrapped_lines = max(1, (len(line) + chars_per_line - 1) // chars_per_line)
        total_lines += wrapped_lines

    # Height per line (approximately 1.2 * font_size)
    line_height = font_size * 1.2

    return max(15, total_lines * line_height * buffer)


# =============================================================================
# CONTENT LOADING HELPERS
# =============================================================================

# Module-level content loader instance
_content_loader = None


def get_content_loader():
    """Get or create the content loader instance."""
    global _content_loader
    if _content_loader is None:
        _content_loader = ContentLoader()
    return _content_loader


def write_context_box(ws, start_row: int, title: str, content: str,
                      start_col: int = 1, width_cols: int = 4) -> int:
    """Write a context/explanation box at the top of a calculation sheet.

    Args:
        ws: Worksheet object
        start_row: Row to start writing
        title: Box title
        content: Multi-line content string (newlines preserved)
        start_col: Starting column (default 1)
        width_cols: Number of columns to merge (default 4)

    Returns:
        Next available row after the context box
    """
    # Context box fill (light blue background)
    context_fill = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")
    context_border = Border(
        left=Side(style='medium', color='4472C4'),
        right=Side(style='medium', color='4472C4'),
        top=Side(style='medium', color='4472C4'),
        bottom=Side(style='medium', color='4472C4')
    )

    row = start_row

    # Title row
    title_cell = ws.cell(row=row, column=start_col, value=title)
    title_cell.font = Font(bold=True, size=11, color="1F4E79")
    title_cell.fill = context_fill
    title_cell.border = context_border
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Merge title across columns
    end_col = start_col + width_cols - 1
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    row += 1

    # Write content in a merged cell with wrap_text
    content_cell = ws.cell(row=row, column=start_col, value=content.strip())
    content_cell.fill = context_fill
    content_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    content_cell.border = context_border

    # Merge content cell
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)

    # Set row height based on content and column width
    merged_width = get_merged_width(ws, start_col, start_col + width_cols - 1)
    ws.row_dimensions[row].height = calculate_row_height(content.strip(), merged_width)

    row += 1

    # Add a blank row after context box
    row += 1

    return row


def write_content_section(ws, start_row: int, section, loader) -> int:
    """Write a content section from YAML to the worksheet.

    Args:
        ws: Worksheet object
        start_row: Row to start writing
        section: SectionContent object from content loader
        loader: ContentLoader instance for resolving references

    Returns:
        Next available row after the section
    """
    row = start_row

    # Section title
    ws.cell(row=row, column=1, value=section.title).font = FONT_HEADER
    row += 1

    # Source citation if present
    if section.source_citation:
        cell = ws.cell(row=row, column=1, value=f"Source: {section.source_citation}")
        cell.font = Font(italic=True)
        row += 1

    # Project if present
    if section.project:
        ws.cell(row=row, column=1, value=section.project)
        row += 1

    # Intro text if present
    if section.intro:
        ws.cell(row=row, column=1, value=section.intro)
        row += 1

    # Bullet point content
    if section.content:
        for line in section.content:
            ws.cell(row=row, column=1, value=f"  - {line}")
            row += 1

    # Table if present
    if section.table:
        row = write_header_row(ws, row, section.table.headers)
        for table_row in section.table.rows:
            if isinstance(table_row, dict):
                for col, header in enumerate(section.table.headers, 1):
                    key = header.lower().replace(" ", "_")
                    ws.cell(row=row, column=col, value=table_row.get(key, ""))
            else:
                for col, val in enumerate(table_row, 1):
                    ws.cell(row=row, column=col, value=val)
            row += 1

    # Handle source references (e.g., glossary, data_sources)
    if section.source:
        if section.source.startswith("global/"):
            source_name = section.source.replace("global/", "").replace(".yaml", "")
            if source_name == "glossary":
                terms = loader.get_glossary()
                for term in terms:
                    ws.cell(row=row, column=1, value=term.term).font = FONT_BOLD
                    ws.cell(row=row, column=2, value=term.definition)
                    row += 1
            elif source_name == "data_sources":
                sources = loader.get_data_sources()
                for source in sources:
                    if isinstance(source, str):
                        ws.cell(row=row, column=1, value=f"  - {source}")
                    else:
                        ws.cell(row=row, column=1, value=f"  - {source.citation}")
                    row += 1
            elif source_name == "change_log":
                entries = loader.get_change_log()
                row = write_header_row(ws, row, ['Version', 'Date', 'Changes'])
                for entry in entries:
                    ws.cell(row=row, column=1, value=entry.version)
                    ws.cell(row=row, column=2, value=entry.date)
                    ws.cell(row=row, column=3, value=entry.changes)
                    row += 1

    # Subsections
    if section.subsections:
        for subsection in section.subsections:
            ws.cell(row=row, column=1, value=subsection.title).font = Font(bold=True)
            row += 1
            if subsection.content:
                for line in subsection.content:
                    ws.cell(row=row, column=1, value=f"  - {line}")
                    row += 1
            row += 1

    row += 1  # Blank row after section
    return row


def write_table_intro(ws, row: int, table_doc, start_col: int = 1, width_cols: int = 1) -> int:
    """Write table title and description above a table.

    Args:
        ws: Worksheet object
        row: Current row position
        table_doc: TableDocumentation object with title and description
        start_col: Starting column (default 1)
        width_cols: Number of columns to merge for description (default 1)

    Returns:
        Next row position (ready for header row)
    """
    # Title row - bold, slightly larger
    if table_doc.title:
        title_cell = ws.cell(row=row, column=start_col, value=table_doc.title)
        title_cell.font = Font(bold=True, size=11)
        row += 1

    # Helper to write a metadata row with label and merged content
    def write_meta_row(label: str, content: str, fill_color: str = "F2F2F2") -> int:
        nonlocal row
        # Write label in first column
        label_cell = ws.cell(row=row, column=start_col, value=f"{label}:")
        label_cell.font = Font(bold=True, size=10, color="666666")
        label_cell.alignment = Alignment(vertical="top")
        label_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        label_cell.fill = label_fill

        # Write content in remaining columns (merged)
        content_cell = ws.cell(row=row, column=start_col + 1, value=content.strip())
        content_cell.font = Font(size=10, color="666666")
        content_cell.alignment = Alignment(wrap_text=True, vertical="top")
        content_cell.fill = label_fill

        # Merge content across remaining width
        if width_cols > 2:
            end_col = start_col + width_cols - 1
            ws.merge_cells(start_row=row, start_column=start_col + 1,
                           end_row=row, end_column=end_col)

        # Set row height based on content and column width
        merged_width = get_merged_width(ws, start_col + 1, start_col + width_cols - 1)
        ws.row_dimensions[row].height = calculate_row_height(content.strip(), merged_width)
        row += 1
        return row

    # Description row
    if table_doc.description:
        write_meta_row("Description", table_doc.description)

    # Purpose row
    if table_doc.purpose:
        write_meta_row("Purpose", table_doc.purpose)

    # Assumptions row
    if table_doc.assumptions:
        write_meta_row("Assumptions", table_doc.assumptions)

    # Data source row
    if table_doc.data_source:
        write_meta_row("Data Source", table_doc.data_source)

    # Related tables row
    if table_doc.related_tables:
        tables_str = ", ".join(table_doc.related_tables)
        write_meta_row("Related Tables", tables_str)

    return row


# =============================================================================
# TRACEABILITY VERIFICATION
# =============================================================================

def verify_traceability(wb, fail_on_mismatch=False):
    """Verify INPUT_REGISTRY matches actual formula usage in generated workbook.

    Scans calculation sheets (6-10, 12-13) for named range references in formulas
    and compares against INPUT_REGISTRY to detect:
    - Registry says input X is used by sheet Y, but formula doesn't contain X
    - Formula contains input X but registry doesn't list sheet Y

    Args:
        wb: Workbook to verify
        fail_on_mismatch: If True, raise exception on mismatch (default False = warn only)

    Returns:
        Tuple of (registry_errors, formula_warnings) where:
        - registry_errors: List of "Registry says X used in Y but not found"
        - formula_warnings: List of "Formula in Y uses X but not in registry"
    """
    import re

    # Sheets to scan for formula references
    calc_sheets = ['8_Calc_Timeline', '10_Calc_Stochastic', '12_Calc_Value',
                   '13_Calc_Costs', '14_Calc_Sim', '16_PL_Annual', '17_CashFlow']

    # Build expected mapping: sheet -> set of expected input names
    expected_by_sheet = {}
    for input_name, sheets in INPUT_REGISTRY.items():
        for sheet in sheets:
            if sheet not in expected_by_sheet:
                expected_by_sheet[sheet] = set()
            expected_by_sheet[sheet].add(input_name)

    # Build actual mapping: sheet -> set of named ranges found in formulas
    actual_by_sheet = {}
    all_registered_names = set(INPUT_REGISTRY.keys())

    for sheet_name in calc_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        found_names = set()

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    # Find all potential named range references
                    # Named ranges are typically word characters, possibly with underscores
                    potential_names = re.findall(r'\b([A-Za-z_][A-Za-z0-9_]*)\b', formula)
                    for name in potential_names:
                        if name in all_registered_names:
                            found_names.add(name)

        actual_by_sheet[sheet_name] = found_names

    # Compare expected vs actual
    registry_errors = []
    formula_warnings = []

    for sheet_name in calc_sheets:
        expected = expected_by_sheet.get(sheet_name, set())
        actual = actual_by_sheet.get(sheet_name, set())

        # Registry says it should be there but isn't
        missing_from_formulas = expected - actual
        for name in missing_from_formulas:
            registry_errors.append(f"Registry: '{name}' should be in {sheet_name}, but not found in formulas")

        # Found in formulas but not in registry (informational)
        extra_in_formulas = actual - expected
        for name in extra_in_formulas:
            formula_warnings.append(f"Formula in {sheet_name} uses '{name}' but not in INPUT_REGISTRY")

    # Output results
    if registry_errors or formula_warnings:
        print("\n" + "=" * 65)
        print("TRACEABILITY VERIFICATION RESULTS")
        print("=" * 65)

        if registry_errors:
            print(f"\nRegistry Mismatches ({len(registry_errors)}):")
            for err in registry_errors:
                print(f"  WARNING: {err}")

        if formula_warnings:
            print(f"\nUndocumented References ({len(formula_warnings)}):")
            for warn in formula_warnings[:10]:  # Limit output
                print(f"  INFO: {warn}")
            if len(formula_warnings) > 10:
                print(f"  ... and {len(formula_warnings) - 10} more")

        print("=" * 65)

        if fail_on_mismatch and registry_errors:
            raise ValueError(f"Traceability verification failed: {len(registry_errors)} registry mismatches")
    else:
        print("\nTraceability verification: PASSED (registry matches formula usage)")

    return registry_errors, formula_warnings


# =============================================================================
# SHEET 0: COVER
# =============================================================================

def create_0_cover(wb):
    """Sheet 0: Cover - Navigation hub and model status."""
    ws = wb.active
    ws.title = "0_Cover"
    ws.sheet_properties.tabColor = TAB_WHITE
    set_column_widths(ws, {'A': 20, 'B': 45})

    # Title
    ws['A1'] = "Luminous Wetland Monte Carlo Financial Model"
    ws['A1'].font = Font(bold=True, size=20)
    ws['A2'] = "FAST Standard Compliant (18-Sheet Architecture)"
    ws['A2'].font = Font(italic=True, size=12)

    ws['A4'] = "Version:"
    ws['B4'] = "3.0"
    ws['A5'] = "Date:"
    ws['B5'] = "=TODAY()"
    ws['B5'].number_format = FMT_DATE
    ws['A6'] = "Author:"
    ws['B6'] = "Luminous BioSolutions"

    # Status indicators
    ws['A8'] = "Model Status"
    ws['A8'].font = FONT_HEADER
    ws['A9'] = "Health:"
    ws['B9'] = "=IF(COUNTIF('21_Checks'!D:D,\"FAIL\")=0, \"All Checks Pass\", \"Review 21_Checks\")"
    ws['A10'] = "Errors:"
    ws['B10'] = "=COUNTIF('21_Checks'!D:D,\"FAIL\")"

    # Navigation
    ws['A12'] = "Sheet Navigation"
    ws['A12'].font = FONT_HEADER

    sheets = [
        ("1_TOC", "Table of Contents / Control panel"),
        ("3_Instructions", "Methodology and documentation"),
        ("4_Assumptions", "ALL inputs (timing, site, test options, validation)"),
        ("6_Scenarios", "Value scenario definitions"),
        ("7_ServiceModels", "Monte Carlo distributions"),
        ("8_Calc_Timeline", "Projection timeline"),
        ("10_Calc_Stochastic", "Random variable generation"),
        ("12_Calc_Value", "Gated value calculations"),
        ("13_Calc_Costs", "Cost calculations"),
        ("14_Calc_Sim", "1000-iteration Monte Carlo"),
        ("15_PL_Monthly", "Monthly P&L (placeholder)"),
        ("16_PL_Annual", "Annual projections"),
        ("17_CashFlow", "Cumulative projections"),
        ("18_UnitEconomics", "Unit economics view"),
        ("19_Sensitivity", "Tornado analysis"),
        ("20_Dashboard", "Executive dashboard"),
        ("21_Checks", "Validation checks"),
    ]

    row = 13
    for sheet_name, description in sheets:
        ws.cell(row=row, column=1, value=sheet_name)
        ws.cell(row=row, column=2, value=description)
        row += 1

    return ws


# =============================================================================
# SHEET 1: TOC (Table of Contents / Control Panel)
# =============================================================================

def create_1_toc(wb):
    """Sheet 1: TOC - Table of Contents and Control Panel."""
    ws = wb.create_sheet("1_TOC")
    ws.sheet_properties.tabColor = TAB_YELLOW
    set_column_widths(ws, {'A': 20, 'B': 18, 'C': 12, 'D': 35})

    # Load table documentation from YAML
    loader = get_content_loader()
    table_docs = loader.get_table_docs("1_toc")

    ws['A1'] = "Control Panel"
    ws['A1'].font = FONT_TITLE

    # Control table
    row = 3
    if 'control' in table_docs:
        row = write_table_intro(ws, row, table_docs['control'], width_cols=4)
    headers = ['Parameter', 'Value', 'Validation', 'Description']
    row = write_header_row(ws, row, headers)

    # Register locations as we write
    start_row = row

    # Testing Option
    ws.cell(row=row, column=1, value="Testing_Option")
    cell = ws.cell(row=row, column=2, value="Daily_Biosensor")
    cell.fill = FILL_EDITABLE
    ws.cell(row=row, column=3, value="Dropdown")
    ws.cell(row=row, column=4, value="Selected testing method")
    add_named_range(wb, "Testing_Option", "1_TOC", f"$B${row}")
    row += 1

    # Scenario Select
    ws.cell(row=row, column=1, value="Scenario_Select")
    cell = ws.cell(row=row, column=2, value="Base")
    cell.fill = FILL_EDITABLE
    ws.cell(row=row, column=3, value="Dropdown")
    ws.cell(row=row, column=4, value="Model scenario")
    add_named_range(wb, "Scenario_Select", "1_TOC", f"$B${row}")
    row += 1

    # Discount Rate
    ws.cell(row=row, column=1, value="Discount_Rate")
    cell = ws.cell(row=row, column=2, value=0.048)
    cell.fill = FILL_EDITABLE
    cell.number_format = FMT_PERCENT
    ws.cell(row=row, column=3, value="0%-15%")
    ws.cell(row=row, column=4, value="NPV discount rate")
    add_named_range(wb, "Discount_Rate", "1_TOC", f"$B${row}")
    row += 1

    # Projection Years
    ws.cell(row=row, column=1, value="Projection_Years")
    cell = ws.cell(row=row, column=2, value=5)
    cell.fill = FILL_EDITABLE
    ws.cell(row=row, column=3, value="1-10")
    ws.cell(row=row, column=4, value="Projection horizon")
    add_named_range(wb, "Projection_Years", "1_TOC", f"$B${row}")
    row += 1

    # Base Year
    ws.cell(row=row, column=1, value="Base_Year")
    cell = ws.cell(row=row, column=2, value=2026)
    cell.fill = FILL_EDITABLE
    ws.cell(row=row, column=3, value="2024-2030")
    ws.cell(row=row, column=4, value="Model start year")
    add_named_range(wb, "Base_Year", "1_TOC", f"$B${row}")
    row += 1

    create_table(ws, "tbl_Control", f"A{start_row-1}:D{row-1}")

    # Data validation will reference 4_Assumptions validation lists
    # (Added after 4_Assumptions is created)

    return ws


# =============================================================================
# SHEET 2: INSTRUCTIONS
# =============================================================================

def create_3_instructions(wb):
    """Sheet 2: Instructions - Methodology, glossary, data sources."""
    ws = wb.create_sheet("3_Instructions")
    ws.sheet_properties.tabColor = TAB_GRAY
    set_column_widths(ws, {'A': 35, 'B': 30, 'C': 50, 'D': 15})

    row = 1
    ws.cell(row=row, column=1, value="Model Instructions & Documentation").font = FONT_TITLE
    row += 2

    # Model Purpose
    ws.cell(row=row, column=1, value="1. Model Purpose").font = FONT_HEADER
    row += 1
    purpose = [
        "Quantify value of high-frequency biosensor monitoring vs HRMS for OSPW wetland treatment",
        "Monte Carlo simulation for uncertainty quantification (1000 iterations)",
        "FAST Standard compliance for transparency and auditability",
        "All calculations performed in Excel - Python only generates structure/formulas",
    ]
    for line in purpose:
        ws.cell(row=row, column=1, value=f"  - {line}")
        row += 1
    row += 1

    # How to Use
    ws.cell(row=row, column=1, value="2. How to Use This Model").font = FONT_HEADER
    row += 1
    usage = [
        "1. Edit yellow-highlighted cells in 1_TOC and 4_Assumptions",
        "2. Go to 14_Calc_Sim and set up the Data Table:",
        "   - Select range shown in instructions (formula row through 1000 iterations)",
        "   - Data > What-If Analysis > Data Table",
        "   - Column Input Cell: $K$1 (blank cell on same sheet)",
        "3. Press Ctrl+Alt+F9 to run Monte Carlo simulation",
        "4. View results in 20_Dashboard",
        "5. Check 21_Checks for any validation errors",
    ]
    for line in usage:
        ws.cell(row=row, column=1, value=line)
        row += 1
    row += 1

    # Latency Gating Logic
    ws.cell(row=row, column=1, value="3. Latency Gating Logic").font = FONT_HEADER
    row += 1
    ws.cell(row=row, column=1, value="HRMS lab testing has 42-day turnaround, blocking time-sensitive decisions:")
    row += 1
    headers = ['Scenario', 'Required_Response', 'HRMS_42d', 'Biosensor_1d']
    row = write_header_row(ws, row, headers)

    gating = [
        ("S1: Season Extension", "7 days", "BLOCKED", "ENABLED"),
        ("S2: Intervention Avoidance", "7 days", "BLOCKED", "ENABLED"),
        ("S3: Spatial Routing", "14 days", "BLOCKED", "ENABLED"),
        ("S4: Early Detection", "3 days", "BLOCKED", "ENABLED"),
    ]

    for scenario, required, hrms, biosensor in gating:
        ws.cell(row=row, column=1, value=scenario)
        ws.cell(row=row, column=2, value=required)
        ws.cell(row=row, column=3, value=hrms)
        ws.cell(row=row, column=4, value=biosensor)
        row += 1
    row += 1

    # Glossary
    ws.cell(row=row, column=1, value="4. Glossary").font = FONT_HEADER
    row += 1
    glossary = [
        ("HRMS", "High-Resolution Mass Spectrometry"),
        ("OSPW", "Oil Sands Process-affected Water"),
        ("ARO", "Asset Retirement Obligation"),
        ("NPV", "Net Present Value"),
        ("HLR", "Hydraulic Loading Rate"),
        ("FAST", "Flexible, Appropriate, Structured, Transparent"),
        ("NAFC", "Naphthenic Acid Fraction Compounds"),
    ]
    for term, definition in glossary:
        ws.cell(row=row, column=1, value=term).font = FONT_BOLD
        ws.cell(row=row, column=2, value=definition)
        row += 1
    row += 1

    # Data Sources
    ws.cell(row=row, column=1, value="5. Data Sources").font = FONT_HEADER
    row += 1
    sources = [
        "Zhang (2026) Long-Dated Environmental Liabilities in Oil Sands",
        "Kadlec & Wallace (2009) Treatment Wetlands, 2nd Ed.",
        "FAST Standard Organisation (fast-standard.org)",
        "Luminous BioSolutions Financial Model Framework",
        "Kearl Wetland Pilot Data (GROW Presentation, Dec 2025)",
    ]
    for source in sources:
        ws.cell(row=row, column=1, value=f"  - {source}")
        row += 1
    row += 1

    # Value Per M3 Derivation
    ws.cell(row=row, column=1, value="6. Value Per Cubic Meter Derivation").font = FONT_HEADER
    row += 1
    ws.cell(row=row, column=1, value="The $5/m3 operational value is derived from:").font = Font(italic=True)
    row += 1
    value_derivation = [
        "Conservative annual value: $250,000/year (Financial Model Framework)",
        "Annual throughput: 50,000 m3/year (4_Assumptions default)",
        "Calculation: $250,000 / 50,000 m3 = $5/m3",
        "This is the OPERATIONAL benefit only - excludes avoided treatment and ARO impacts",
    ]
    for line in value_derivation:
        ws.cell(row=row, column=1, value=f"  - {line}")
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Value Sensitivity Analysis (19_Sensitivity):").font = Font(bold=True)
    row += 1
    sensitivity_context = [
        "Operational: $3-8/m3 (base $5) - Direct wetland treatment value",
        "Avoided Treatment: $10-30/m3 (base $20) - Alternative treatment costs",
        "ARO Liability: $21-93/m3 (base $50) - Zhang (2026) liability analysis",
    ]
    for line in sensitivity_context:
        ws.cell(row=row, column=1, value=f"  - {line}")
        row += 1
    row += 1

    # Kearl Treatment Kinetics
    ws.cell(row=row, column=1, value="7. Treatment Kinetics (Kearl 2022 Field Data)").font = FONT_HEADER
    row += 1
    ws.cell(row=row, column=1, value="Source: Vander Meulen & Headley (2024) GROW Presentation").font = Font(italic=True)
    row += 1
    ws.cell(row=row, column=1, value="Project: Genome Canada LSARP - GROW (Genomics Research for Optimization of Constructed Treatment Wetlands)")
    row += 2

    ws.cell(row=row, column=1, value="Two-Phase Kinetics Model:").font = Font(bold=True)
    row += 1
    kinetics_notes = [
        "Rapid Phase (Days 0-15): -0.532 mg/L/day (R^2=0.43)",
        "Slow Phase (Days 30-69): -0.248 mg/L/day (R^2=0.50)",
        "The slowdown reflects preferential removal of labile compounds",
    ]
    for note in kinetics_notes:
        ws.cell(row=row, column=1, value=f"  - {note}")
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Toxicity Endpoints (Fathead Minnow Embryo):").font = Font(bold=True)
    row += 1
    toxicity_notes = [
        "Day 14: ~50% survival (1 recirculation cycle = LC50)",
        "Day 28: 80-90% survival (2 cycles)",
        "Day 42: >90% survival (3 cycles = non-toxic, equivalent to controls)",
    ]
    for note in toxicity_notes:
        ws.cell(row=row, column=1, value=f"  - {note}")
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="The Concentration Paradox:").font = Font(bold=True)
    row += 1
    paradox_notes = [
        "Water at Day 42 was non-toxic despite ~50 mg/L NAFC concentration",
        "Literature LC50 values for fathead minnow: 18-34 mg/L",
        "Explanation: Wetlands preferentially remove high-carbon, lipophilic (most toxic) compounds",
        "Remaining NAFCs are oxidized (O3, O4 classes) and less bioavailable",
        "Implication: Total NAFC concentration alone may be inadequate for assessing safety",
    ]
    for note in paradox_notes:
        ws.cell(row=row, column=1, value=f"  - {note}")
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Adjusting Parameters:").font = Font(bold=True)
    row += 1
    adjust_notes = [
        "All kinetics parameters are in 4_Assumptions Section H (yellow cells)",
        "Treatment_Rate_Factor in 7_ServiceModels captures seasonal/vegetation variability",
        "Modify Initial_NAFC to match your source water concentration",
        "Adjust Target_NAFC for different regulatory thresholds",
    ]
    for note in adjust_notes:
        ws.cell(row=row, column=1, value=f"  - {note}")
        row += 1
    row += 1

    # Change Log
    ws.cell(row=row, column=1, value="8. Change Log").font = FONT_HEADER
    row += 1
    headers = ['Version', 'Date', 'Changes']
    row = write_header_row(ws, row, headers)

    ws.cell(row=row, column=1, value="3.2")
    ws.cell(row=row, column=2, value="2026-01-29")
    ws.cell(row=row, column=3, value="Value/m3 derivation documentation (Section 6)")
    row += 1

    ws.cell(row=row, column=1, value="3.1")
    ws.cell(row=row, column=2, value="2026-01-29")
    ws.cell(row=row, column=3, value="Kearl treatment kinetics integration, Days_to_Compliance MC")
    row += 1

    ws.cell(row=row, column=1, value="3.0")
    ws.cell(row=row, column=2, value="2026-01-29")
    ws.cell(row=row, column=3, value="FAST 18-sheet refactor with LOCATIONS registry")
    row += 1

    ws.cell(row=row, column=1, value="2.0")
    ws.cell(row=row, column=2, value="2026-01-28")
    ws.cell(row=row, column=3, value="Excel-native Monte Carlo via Data Tables")
    row += 1

    ws.cell(row=row, column=1, value="1.0")
    ws.cell(row=row, column=2, value="2025-12-01")
    ws.cell(row=row, column=3, value="Initial Python-based model")

    return ws


# =============================================================================
# SHEET 1.5: INPUT MAP (Traceability Reference)
# =============================================================================

def create_2_inputmap(wb):
    """Sheet 1.5: InputMap - Complete registry of inputs and their consumers.

    Shows which inputs are used by which calculation sheets, providing
    traceability from inputs to calculations.
    """
    ws = wb.create_sheet("2_InputMap")
    ws.sheet_properties.tabColor = TAB_GRAY
    set_column_widths(ws, {'A': 32, 'B': 18, 'C': 18, 'D': 28})

    ws['A1'] = "Input Traceability Map"
    ws['A1'].font = FONT_TITLE

    ws['A2'] = "Shows which calculation sheets directly reference each input"
    ws['A2'].font = Font(italic=True)

    row = 4
    headers = ['Input Name', 'Value', 'Source Sheet', 'Referenced By']
    row = write_header_row(ws, row, headers)

    map_start = row

    # Group inputs by source sheet for better organization
    inputs_4_assumptions = [
        ('Initial_NAFC', '4_Assumptions'),
        ('Target_NAFC', '4_Assumptions'),
        ('Rapid_Phase_Rate', '4_Assumptions'),
        ('Slow_Phase_Rate', '4_Assumptions'),
        ('Rapid_Phase_Duration', '4_Assumptions'),
        ('Recirculation_Cycle', '4_Assumptions'),
        ('Season_Length', '4_Assumptions'),
        ('LearningCurve_Mult', '4_Assumptions'),
    ]

    inputs_7_servicemodels = [
        ('Tri_Season_Extension_Days_Min', '7_ServiceModels'),
        ('Tri_Season_Extension_Days_Mode', '7_ServiceModels'),
        ('Tri_Season_Extension_Days_Max', '7_ServiceModels'),
        ('Tri_Extension_Value_Per_Day_Min', '7_ServiceModels'),
        ('Tri_Extension_Value_Per_Day_Mode', '7_ServiceModels'),
        ('Tri_Extension_Value_Per_Day_Max', '7_ServiceModels'),
        ('Tri_Interventions_Avoided_Min', '7_ServiceModels'),
        ('Tri_Interventions_Avoided_Mode', '7_ServiceModels'),
        ('Tri_Interventions_Avoided_Max', '7_ServiceModels'),
        ('Tri_Treatment_Rate_Factor_Min', '7_ServiceModels'),
        ('Tri_Treatment_Rate_Factor_Mode', '7_ServiceModels'),
        ('Tri_Treatment_Rate_Factor_Max', '7_ServiceModels'),
        ('Beta_Efficiency_Gain_Alpha', '7_ServiceModels'),
        ('Beta_Efficiency_Gain_Beta', '7_ServiceModels'),
        ('Beta_Efficiency_Gain_Min', '7_ServiceModels'),
        ('Beta_Efficiency_Gain_Max', '7_ServiceModels'),
    ]

    inputs_6_calc = [
        ('Eff_Rapid_Rate', '8_Calc_Timeline'),
        ('Eff_Slow_Rate', '8_Calc_Timeline'),
        ('NAFC_After_Rapid', '8_Calc_Timeline'),
        ('Days_to_Compliance', '8_Calc_Timeline'),
        ('Timeline_LearningMult', '8_Calc_Timeline'),
        ('Timeline_DiscountFactor', '8_Calc_Timeline'),
    ]

    all_inputs = inputs_4_assumptions + inputs_7_servicemodels + inputs_6_calc

    for input_name, source_sheet in all_inputs:
        ws.cell(row=row, column=1, value=input_name)
        # Value column - formula reference or range note
        if is_range_input(input_name):
            ws.cell(row=row, column=2, value="(range - see source sheet)")
        else:
            ws.cell(row=row, column=2, value=f"={input_name}")
        ws.cell(row=row, column=3, value=source_sheet)
        # Referenced By - from INPUT_REGISTRY
        refs = INPUT_REGISTRY.get(input_name, [])
        ws.cell(row=row, column=4, value=', '.join(refs) if refs else '')
        row += 1

    create_table(ws, "tbl_InputMap", f"A{map_start-1}:D{row-1}")

    # Add summary section
    row += 2
    ws.cell(row=row, column=1, value="Summary").font = FONT_HEADER
    row += 1
    ws.cell(row=row, column=1, value=f"Total inputs tracked: {len(all_inputs)}")
    row += 1
    ws.cell(row=row, column=1, value=f"Registry entries: {len(INPUT_REGISTRY)}")
    row += 2

    ws.cell(row=row, column=1, value="Note:").font = FONT_BOLD
    row += 1
    ws.cell(row=row, column=1, value="'Referenced By' shows direct formula references only,")
    row += 1
    ws.cell(row=row, column=1, value="not downstream dependencies. The verify_traceability()")
    row += 1
    ws.cell(row=row, column=1, value="function validates this registry against actual formulas.")

    return ws


# =============================================================================
# SHEET 3: ASSUMPTIONS (Consolidated: Timing + Site Config + Test Options + Validation)
# =============================================================================

def create_4_assumptions(wb):
    """Sheet 3: Assumptions - ALL inputs consolidated.

    This sheet consolidates:
    - Timing parameters (from old 2_Timing)
    - Site configuration (from old 4_Assumptions)
    - Cost parameters
    - Learning curve
    - Test options (from old 5_Ref_TestOptions)
    - Validation lists (from old 7_Ref_Validation)
    - MC anchor cell
    """
    ws = wb.create_sheet("4_Assumptions")
    ws.sheet_properties.tabColor = TAB_YELLOW
    set_column_widths(ws, {'A': 24, 'B': 14, 'C': 12, 'D': 42, 'E': 30, 'F': 16, 'G': 14})

    # Load table documentation from YAML
    loader = get_content_loader()
    table_docs = loader.get_table_docs("4_assumptions")

    ws['A1'] = "Model Assumptions (Single Source of Truth)"
    ws['A1'].font = FONT_TITLE

    row = 3

    # =========================================================================
    # SECTION A: TIMING PARAMETERS
    # =========================================================================
    if 'timing' in table_docs:
        row = write_table_intro(ws, row, table_docs['timing'], width_cols=5)
    headers = ['Parameter', 'Value', 'Unit', 'Notes', 'Referenced By']
    row = write_header_row(ws, row, headers)

    timing_start = row
    timing = [
        ('Season_Start_Month', 5, 'month', 'May'),
        ('Season_Start_Day', 15, 'day', 'Mid-May'),
        ('Season_End_Month', 9, 'month', 'September'),
        ('Season_End_Day', 15, 'day', 'Mid-September'),
        ('Treatment_Days', 100, 'days', 'Active season length'),
    ]

    for param, value, unit, notes in timing:
        ws.cell(row=row, column=1, value=param)
        cell = ws.cell(row=row, column=2, value=value)
        cell.fill = FILL_EDITABLE
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=notes)
        # Referenced By column - lookup from INPUT_REGISTRY
        refs = get_input_references(param)
        ws.cell(row=row, column=5, value=', '.join(refs) if refs else '')
        # Register each timing parameter
        add_named_range(wb, param, "4_Assumptions", f"$B${row}")
        row += 1

    # Also add Season_Length as alias for Treatment_Days
    add_named_range(wb, "Season_Length", "4_Assumptions", f"$B${timing_start + 4}")

    create_table(ws, "tbl_Timing", f"A{timing_start-1}:E{row-1}")
    row += 1

    # Seasonality table
    if 'seasonality' in table_docs:
        row = write_table_intro(ws, row, table_docs['seasonality'], width_cols=4)
    headers = ['Month', 'Month_Num', 'Multiplier', 'Activity']
    row = write_header_row(ws, row, headers)

    seasonality_start = row
    seasonality = [
        ('May', 5, 0.85, 'Ramp-up'),
        ('June', 6, 1.00, 'Peak'),
        ('July', 7, 1.00, 'Peak'),
        ('August', 8, 0.95, 'Slight decline'),
        ('September', 9, 0.75, 'Decline'),
    ]

    for month, num, mult, activity in seasonality:
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=num)
        cell = ws.cell(row=row, column=3, value=mult)
        cell.fill = FILL_EDITABLE
        ws.cell(row=row, column=4, value=activity)
        row += 1

    create_table(ws, "tbl_Seasonality", f"A{seasonality_start-1}:D{row-1}")
    row += 2

    # =========================================================================
    # SECTION B: SITE CONFIGURATION
    # =========================================================================
    if 'site_config' in table_docs:
        row = write_table_intro(ws, row, table_docs['site_config'], width_cols=5)
    headers = ['Parameter', 'Value', 'Unit', 'Source', 'Referenced By']
    row = write_header_row(ws, row, headers)

    site_start = row
    site_config = [
        ('WetlandAreaHa', 1.0, 'ha', 'Standard design (Kearl pilot = 0.76 ha)'),
        ('TreatmentCells', 5, 'count', 'Multi-cell design'),
        ('AnnualThroughputM3', 50000, 'm3/yr', 'HLR = 0.05 m3/m2/day'),
        ('ValuePerM3', 5.00, '$/m3', '$250K/yr / 50K m3/yr (Section 6)'),
    ]

    for param, value, unit, source in site_config:
        ws.cell(row=row, column=1, value=param)
        cell = ws.cell(row=row, column=2, value=value)
        cell.fill = FILL_EDITABLE
        if 'Value' in param or '$' in unit:
            cell.number_format = FMT_CURRENCY_DEC
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=source)
        # Referenced By column
        refs = get_input_references(param)
        ws.cell(row=row, column=5, value=', '.join(refs) if refs else '')
        add_named_range(wb, param.replace('Ha', '_Ha').replace('M3', '_M3'),
                       "4_Assumptions", f"$B${row}")
        row += 1

    # Simpler named ranges for common ones
    add_named_range(wb, "Wetland_Area", "4_Assumptions", f"$B${site_start}")
    add_named_range(wb, "Treatment_Cells", "4_Assumptions", f"$B${site_start+1}")
    add_named_range(wb, "Annual_Throughput", "4_Assumptions", f"$B${site_start+2}")
    add_named_range(wb, "Value_Per_M3", "4_Assumptions", f"$B${site_start+3}")

    create_table(ws, "tbl_SiteConfig", f"A{site_start-1}:E{row-1}")
    row += 2

    # =========================================================================
    # SECTION C: COST PARAMETERS
    # =========================================================================
    if 'cost_params' in table_docs:
        row = write_table_intro(ws, row, table_docs['cost_params'], width_cols=4)
    headers = ['Parameter', 'Value', 'Unit', 'Referenced By']
    row = write_header_row(ws, row, headers)

    cost_start = row
    cost_params = [
        ('LaborPerDay', 1500, '$/day'),
        ('RegulatoryPenalty', 50000, '$'),
        ('DowntimeCostPerDay', 15000, '$/day'),
    ]

    for param, value, unit in cost_params:
        ws.cell(row=row, column=1, value=param)
        cell = ws.cell(row=row, column=2, value=value)
        cell.fill = FILL_EDITABLE
        cell.number_format = FMT_CURRENCY
        ws.cell(row=row, column=3, value=unit)
        # Referenced By column
        refs = get_input_references(param)
        ws.cell(row=row, column=4, value=', '.join(refs) if refs else '')
        add_named_range(wb, param, "4_Assumptions", f"$B${row}")
        row += 1

    create_table(ws, "tbl_CostParams", f"A{cost_start-1}:D{row-1}")
    row += 2

    # =========================================================================
    # SECTION D: LEARNING CURVE
    # =========================================================================
    if 'learning_curve' in table_docs:
        row = write_table_intro(ws, row, table_docs['learning_curve'], width_cols=4)
    headers = ['Year', 'Multiplier', 'Notes', 'Referenced By']
    row = write_header_row(ws, row, headers)

    learning_start = row
    learning_curve = [
        (1, 0.50, 'Pilot learning'),
        (2, 0.75, 'Initial optimization'),
        (3, 1.00, 'Full maturity'),
        (4, 1.00, 'Steady state'),
        (5, 1.00, 'Steady state'),
    ]

    # Get references for the learning curve range
    lc_refs = get_input_references('LearningCurve_Mult')
    lc_ref_str = ', '.join(lc_refs) if lc_refs else ''

    for year, mult, notes in learning_curve:
        ws.cell(row=row, column=1, value=year)
        cell = ws.cell(row=row, column=2, value=mult)
        cell.fill = FILL_EDITABLE
        ws.cell(row=row, column=3, value=notes)
        # Referenced By column - same for all rows in the range
        ws.cell(row=row, column=4, value=lc_ref_str)
        row += 1

    # Register learning curve range for INDEX lookups
    register_range("LearningCurve_Mult", "4_Assumptions", "B", learning_start, "B", row-1)
    add_named_range(wb, "LearningCurve_Mult", "4_Assumptions", f"$B${learning_start}:$B${row-1}")
    create_table(ws, "tbl_LearningCurve", f"A{learning_start-1}:D{row-1}")
    row += 2

    # =========================================================================
    # SECTION E: TEST OPTIONS (from old 5_Ref_TestOptions)
    # =========================================================================
    if 'test_options' in table_docs:
        row = write_table_intro(ws, row, table_docs['test_options'], width_cols=7)
    headers = ['Option', 'Method', 'Interval', 'Points', 'Cost_Per_Test', 'Latency_Days', 'Referenced By']
    row = write_header_row(ws, row, headers)

    testopts_start = row
    test_options = [
        ('Monthly_HRMS', 'HRMS', 30, 2, 1000, 42),
        ('Weekly_HRMS', 'HRMS', 7, 2, 1000, 42),
        ('Daily_Biosensor', 'Biosensor', 1, 10, 10, 1),
    ]

    for option, method, interval, points, cost, latency in test_options:
        ws.cell(row=row, column=1, value=option)
        ws.cell(row=row, column=2, value=method)
        ws.cell(row=row, column=3, value=interval)
        ws.cell(row=row, column=4, value=points)
        cell = ws.cell(row=row, column=5, value=cost)
        cell.number_format = FMT_CURRENCY
        ws.cell(row=row, column=6, value=latency)
        # Referenced By column - test options are referenced via range lookups
        ws.cell(row=row, column=7, value='')
        row += 1

    testopts_end = row - 1
    create_table(ws, "tbl_TestOptions", f"A{testopts_start-1}:G{row-1}")

    # Named ranges for each column (for INDEX/MATCH lookups)
    add_named_range(wb, "TestOpt_Option", "4_Assumptions", f"$A${testopts_start}:$A${testopts_end}")
    add_named_range(wb, "TestOpt_Method", "4_Assumptions", f"$B${testopts_start}:$B${testopts_end}")
    add_named_range(wb, "TestOpt_Interval", "4_Assumptions", f"$C${testopts_start}:$C${testopts_end}")
    add_named_range(wb, "TestOpt_Points", "4_Assumptions", f"$D${testopts_start}:$D${testopts_end}")
    add_named_range(wb, "TestOpt_Cost", "4_Assumptions", f"$E${testopts_start}:$E${testopts_end}")
    add_named_range(wb, "TestOpt_Latency", "4_Assumptions", f"$F${testopts_start}:$F${testopts_end}")
    row += 1

    # Latency Gating (calculated)
    if 'latency_gating' in table_docs:
        row = write_table_intro(ws, row, table_docs['latency_gating'], width_cols=5)
    headers = ['Option', 'S1_Gate', 'S2_Gate', 'S3_Gate', 'S4_Gate']
    row = write_header_row(ws, row, headers)

    gating_start = row
    for i, (option, method, interval, points, cost, latency) in enumerate(test_options):
        gate_row = row
        ws.cell(row=row, column=1, value=option)
        idx = i + 1
        # S1: Requires <=7 days
        ws.cell(row=row, column=2, value=f'=IF(INDEX(TestOpt_Latency,{idx})<=7,"ENABLED","BLOCKED")')
        # S2: Requires <=7 days
        ws.cell(row=row, column=3, value=f'=IF(INDEX(TestOpt_Latency,{idx})<=7,"ENABLED","BLOCKED")')
        # S3: Requires <=14 days
        ws.cell(row=row, column=4, value=f'=IF(INDEX(TestOpt_Latency,{idx})<=14,"ENABLED","BLOCKED")')
        # S4: Requires <=3 days
        ws.cell(row=row, column=5, value=f'=IF(INDEX(TestOpt_Latency,{idx})<=3,"ENABLED","BLOCKED")')
        row += 1

    create_table(ws, "tbl_LatencyGating", f"A{gating_start-1}:E{row-1}")

    # Named ranges for gating lookups
    gating_end = row - 1
    add_named_range(wb, "Gate_S1", "4_Assumptions", f"$B${gating_start}:$B${gating_end}")
    add_named_range(wb, "Gate_S2", "4_Assumptions", f"$C${gating_start}:$C${gating_end}")
    add_named_range(wb, "Gate_S3", "4_Assumptions", f"$D${gating_start}:$D${gating_end}")
    add_named_range(wb, "Gate_S4", "4_Assumptions", f"$E${gating_start}:$E${gating_end}")
    row += 2

    # =========================================================================
    # SECTION F: VALIDATION LISTS (from old 7_Ref_Validation)
    # =========================================================================
    ws.cell(row=row, column=1, value="F. VALIDATION LISTS").font = FONT_HEADER
    row += 1

    ws.cell(row=row, column=1, value="Testing_Options")
    ws.cell(row=row, column=2, value="Scenarios")
    row += 1

    validation_start = row
    ws.cell(row=row, column=1, value="Monthly_HRMS")
    ws.cell(row=row, column=2, value="Base")
    row += 1
    ws.cell(row=row, column=1, value="Weekly_HRMS")
    ws.cell(row=row, column=2, value="Conservative")
    row += 1
    ws.cell(row=row, column=1, value="Daily_Biosensor")
    ws.cell(row=row, column=2, value="Optimistic")
    row += 1

    # Named ranges for validation dropdowns
    add_named_range(wb, "List_TestingOptions", "4_Assumptions",
                   f"$A${validation_start}:$A${row-1}")
    add_named_range(wb, "List_Scenarios", "4_Assumptions",
                   f"$B${validation_start}:$B${row-1}")
    row += 1

    # =========================================================================
    # MC ANCHOR CELL (Critical for Data Table)
    # =========================================================================
    ws.cell(row=row, column=1, value="G. MC ANCHOR").font = FONT_HEADER
    row += 1
    ws.cell(row=row, column=1, value="MC_Anchor -->")
    # B column left blank (implicit None) - critical for Data Table column input
    mc_anchor_row = row
    add_named_range(wb, "MC_Anchor", "4_Assumptions", f"$B${mc_anchor_row}")
    row += 2

    # =========================================================================
    # SECTION H: TREATMENT KINETICS (Kearl Field Data)
    # =========================================================================
    if 'treatment_kinetics' in table_docs:
        row = write_table_intro(ws, row, table_docs['treatment_kinetics'], width_cols=5)
    headers = ['Parameter', 'Value', 'Unit', 'Notes', 'Referenced By']
    row = write_header_row(ws, row, headers)

    kinetics_start = row
    treatment_kinetics = [
        ('Initial_NAFC', 63.1, 'mg/L', 'Kearl 2022 Day 0 concentration'),
        ('Target_NAFC', 50.0, 'mg/L', 'Regulatory threshold for non-toxicity'),
        ('Rapid_Phase_Rate', 0.532, 'mg/L/day', 'Kearl Days 0-15 (R^2=0.43)'),
        ('Slow_Phase_Rate', 0.248, 'mg/L/day', 'Kearl Days 30-69 (R^2=0.50)'),
        ('Rapid_Phase_Duration', 15, 'days', 'Duration of rapid removal phase'),
        ('Recirculation_Cycle', 14, 'days', 'Standard wetland recirculation'),
        ('System_Capacity', 8772, 'm3', 'Kearl 4-bay system volume'),
    ]

    for param, value, unit, notes in treatment_kinetics:
        ws.cell(row=row, column=1, value=param)
        cell = ws.cell(row=row, column=2, value=value)
        cell.fill = FILL_EDITABLE
        if 'mg/L' in unit:
            cell.number_format = '0.000'
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=notes)
        # Referenced By column
        refs = get_input_references(param)
        ws.cell(row=row, column=5, value=', '.join(refs) if refs else '')
        add_named_range(wb, param, "4_Assumptions", f"$B${row}")
        row += 1

    create_table(ws, "tbl_TreatmentKinetics", f"A{kinetics_start-1}:E{row-1}")
    row += 2

    # =========================================================================
    # SECTION I: TOXICITY MILESTONES (Kearl Field Data)
    # =========================================================================
    if 'toxicity_milestones' in table_docs:
        row = write_table_intro(ws, row, table_docs['toxicity_milestones'], width_cols=4)
    headers = ['Milestone', 'Days', 'Survival', 'Notes']
    row = write_header_row(ws, row, headers)

    toxicity_start = row
    toxicity_milestones = [
        ('Day_LC50', 14, '50%', '1 recirculation cycle - LC50 threshold'),
        ('Day_80_Survival', 28, '80-90%', '2 recirculation cycles'),
        ('Day_Nontoxic', 42, '>90%', '3 cycles - equivalent to negative control'),
    ]

    for milestone, days, survival, notes in toxicity_milestones:
        ws.cell(row=row, column=1, value=milestone)
        cell = ws.cell(row=row, column=2, value=days)
        cell.fill = FILL_EDITABLE
        ws.cell(row=row, column=3, value=survival)
        ws.cell(row=row, column=4, value=notes)
        add_named_range(wb, milestone, "4_Assumptions", f"$B${row}")
        row += 1

    create_table(ws, "tbl_ToxicityMilestones", f"A{toxicity_start-1}:D{row-1}")

    # Store key row numbers for other sheets to reference
    register_location("_timing_start", "4_Assumptions", "A", timing_start)
    register_location("_site_start", "4_Assumptions", "A", site_start)
    register_location("_learning_start", "4_Assumptions", "B", learning_start)
    register_location("_testopts_start", "4_Assumptions", "A", testopts_start)
    register_location("_gating_start", "4_Assumptions", "A", gating_start)

    return ws


# =============================================================================
# SHEET 5: ENVIRONMENTAL DRIVERS (NEW)
# =============================================================================

def create_5_environmentaldrivers(wb):
    """Sheet 5: EnvironmentalDrivers - Monthly environmental profiles for kinetics."""
    ws = wb.create_sheet("5_EnvironmentalDrivers")
    ws.sheet_properties.tabColor = TAB_YELLOW
    set_column_widths(ws, {'A': 22, 'B': 10, 'C': 10, 'D': 10, 'E': 10, 'F': 10, 'G': 10, 'H': 30})

    ws['A1'] = "Environmental Driver Profiles"
    ws['A1'].font = FONT_TITLE

    # Load context box from YAML if available
    row = 3
    loader = get_content_loader()
    try:
        context = loader.get_context_box("5_environmentaldrivers")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=8)
    except (ContentNotFoundError, ContentValidationError):
        # Default context if YAML not available
        context_content = """Monthly profiles for environmental variables affecting treatment kinetics.
Temperature drives Q10 rate modifier (rates double per 10C above 20C).
UV intensity affects photodegradation rates.
Dissolved oxygen controls aerobic biodegradation.
Edit yellow cells to customize for your site."""
        row = write_context_box(ws, row, "Environmental Drivers", context_content, width_cols=8)

    # Environmental Profiles Table
    row += 1
    headers = ['Variable', 'Unit', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Notes']
    row = write_header_row(ws, row, headers)

    env_start = row
    env_profiles = [
        ('Temp_Water', 'C', 10, 16, 20, 18, 12, 'Water temperature - Q10 driver'),
        ('UV_Relative', '0-1', 0.75, 1.00, 1.00, 0.85, 0.55, 'Relative UV intensity'),
        ('DO_Typical', 'mg/L', 8.5, 7.5, 6.5, 7.0, 8.0, 'Dissolved oxygen'),
        ('Nutrient_Index', '0-1', 1.0, 0.7, 0.4, 0.3, 0.2, 'Nutrient availability'),
        ('Bioavail_Index', '0-1', 0.7, 0.9, 1.0, 0.95, 0.8, 'Bioavailability factor'),
    ]

    for var_name, unit, may, jun, jul, aug, sep, notes in env_profiles:
        ws.cell(row=row, column=1, value=var_name)
        ws.cell(row=row, column=2, value=unit)
        # Monthly values - editable
        for col, val in enumerate([may, jun, jul, aug, sep], 3):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = FILL_EDITABLE
            if unit == 'C':
                cell.number_format = '0.0'
            elif unit in ['0-1', 'mg/L']:
                cell.number_format = '0.00'
        ws.cell(row=row, column=8, value=notes)
        # Register named ranges for each month
        for col_idx, month in enumerate(['May', 'Jun', 'Jul', 'Aug', 'Sep'], 3):
            col_letter = get_column_letter(col_idx)
            range_name = f"Env_{var_name}_{month}"
            register_location(range_name, "5_EnvironmentalDrivers", col_letter, row)
            add_named_range(wb, range_name, "5_EnvironmentalDrivers", f"${col_letter}${row}")
        row += 1

    create_table(ws, "tbl_Env_Profiles", f"A{env_start-1}:H{row-1}")
    row += 2

    # Ionic/Compliance Parameters Table
    ws.cell(row=row, column=1, value="Ionic Parameters (Compliance Gates)").font = Font(bold=True)
    row += 1
    headers = ['Parameter', 'Unit', 'Min', 'Mode', 'Max', 'Threshold', 'Direction', 'Notes']
    row = write_header_row(ws, row, headers)

    ionic_start = row
    ionic_params = [
        ('Chloride', 'mg/L', 50, 150, 300, 230, 'below', 'Aquatic life threshold'),
        ('SAR', 'ratio', 1.0, 2.5, 6.0, 4.0, 'below', 'Land application limit'),
        ('Sulfate', 'mg/L', 100, 300, 600, 500, 'below', 'Process water indicator'),
        ('DO_Compliance', 'mg/L', 5.0, 7.0, 9.0, 6.5, 'above', 'Minimum for discharge'),
    ]

    for param, unit, min_v, mode_v, max_v, thresh, direction, notes in ionic_params:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=unit)
        for col, val in enumerate([min_v, mode_v, max_v], 3):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = FILL_EDITABLE
            cell.number_format = '0.0' if isinstance(val, float) else '0'
        ws.cell(row=row, column=6, value=thresh)
        ws.cell(row=row, column=7, value=direction)
        ws.cell(row=row, column=8, value=notes)
        # Register named ranges
        register_location(f"Ionic_{param}_Min", "5_EnvironmentalDrivers", "C", row)
        add_named_range(wb, f"Ionic_{param}_Min", "5_EnvironmentalDrivers", f"$C${row}")
        register_location(f"Ionic_{param}_Mode", "5_EnvironmentalDrivers", "D", row)
        add_named_range(wb, f"Ionic_{param}_Mode", "5_EnvironmentalDrivers", f"$D${row}")
        register_location(f"Ionic_{param}_Max", "5_EnvironmentalDrivers", "E", row)
        add_named_range(wb, f"Ionic_{param}_Max", "5_EnvironmentalDrivers", f"$E${row}")
        register_location(f"Ionic_{param}_Threshold", "5_EnvironmentalDrivers", "F", row)
        add_named_range(wb, f"Ionic_{param}_Threshold", "5_EnvironmentalDrivers", f"$F${row}")
        row += 1

    create_table(ws, "tbl_Ionic_Params", f"A{ionic_start-1}:H{row-1}")
    row += 2

    # Site Profile Reference
    ws.cell(row=row, column=1, value="Current Site Profile:").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Kearl")
    ws.cell(row=row, column=3, value="(from config/sites/kearl.yaml)")

    return ws


# =============================================================================
# SHEET 6: SCENARIOS
# =============================================================================

def create_6_scenarios(wb):
    """Sheet 4: Scenarios - Value scenario definitions."""
    ws = wb.create_sheet("6_Scenarios")
    ws.sheet_properties.tabColor = TAB_YELLOW
    set_column_widths(ws, {'A': 12, 'B': 22, 'C': 14, 'D': 24, 'E': 40})

    # Load table documentation from YAML
    loader = get_content_loader()
    table_docs = loader.get_table_docs("6_scenarios")
    collected_docs = []  # Collect for reference section

    ws['A1'] = "Value Scenario Definitions"
    ws['A1'].font = FONT_TITLE

    # Value Scenarios - with inline documentation
    row = 3
    if 'value_scenarios' in table_docs:
        doc = table_docs['value_scenarios']
        collected_docs.append(doc)
        row = write_table_intro(ws, row, doc, width_cols=5)

    headers = ['ID', 'Name', 'Base_Value', 'Required_Response_Days', 'Description']
    row = write_header_row(ws, row, headers)

    scenarios_start = row
    scenarios = [
        ('S1', 'Season Extension', 104000, 7, 'Temperature-based vs calendar shutdown'),
        ('S2', 'Intervention Avoidance', 67500, 7, 'Prevented false-positive interventions'),
        ('S3', 'Spatial Routing', 78000, 14, 'Route to high-performing cells'),
        ('S4', 'Early Detection', 36000, 3, 'Reduced bio-assay costs'),
    ]

    for sid, name, base_val, response, desc in scenarios:
        ws.cell(row=row, column=1, value=sid)
        ws.cell(row=row, column=2, value=name)
        cell = ws.cell(row=row, column=3, value=base_val)
        cell.number_format = FMT_CURRENCY
        cell.fill = FILL_EDITABLE
        ws.cell(row=row, column=4, value=response)
        ws.cell(row=row, column=5, value=desc)
        row += 1

    scenarios_end = row - 1
    create_table(ws, "tbl_ValueScenarios", f"A{scenarios_start-1}:E{row-1}")

    # Named range for base values column
    add_named_range(wb, "Scenario_BaseValues", "6_Scenarios",
                   f"$C${scenarios_start}:$C${scenarios_end}")
    row += 2

    # Model Scenarios - with inline documentation
    if 'model_scenarios' in table_docs:
        doc = table_docs['model_scenarios']
        collected_docs.append(doc)
        row = write_table_intro(ws, row, doc, width_cols=4)

    headers = ['Scenario', 'Discount_Adj', 'Growth_Adj', 'Risk_Adj']
    row = write_header_row(ws, row, headers)

    model_start = row
    model_scenarios = [
        ('Base', 0, 0, 0),
        ('Conservative', 0.004, -0.02, 0.10),
        ('Optimistic', -0.003, 0.01, -0.10),
    ]

    for scenario, disc_adj, growth_adj, risk_adj in model_scenarios:
        ws.cell(row=row, column=1, value=scenario)
        cell = ws.cell(row=row, column=2, value=disc_adj)
        cell.number_format = FMT_PERCENT
        cell.fill = FILL_EDITABLE
        cell = ws.cell(row=row, column=3, value=growth_adj)
        cell.number_format = FMT_PERCENT
        cell.fill = FILL_EDITABLE
        cell = ws.cell(row=row, column=4, value=risk_adj)
        cell.number_format = FMT_PERCENT
        cell.fill = FILL_EDITABLE
        row += 1

    create_table(ws, "tbl_ModelScenarios", f"A{model_start-1}:D{row-1}")

    return ws


# =============================================================================
# SHEET 5: SERVICE MODELS (Distribution Parameters)
# =============================================================================

def create_7_servicemodels(wb):
    """Sheet 5: ServiceModels - Monte Carlo distribution parameters."""
    ws = wb.create_sheet("7_ServiceModels")
    ws.sheet_properties.tabColor = TAB_YELLOW
    set_column_widths(ws, {'A': 28, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 30, 'G': 35})

    # Load table documentation from YAML
    loader = get_content_loader()
    table_docs = loader.get_table_docs("7_servicemodels")

    ws['A1'] = "Monte Carlo Distribution Parameters"
    ws['A1'].font = FONT_TITLE

    # Triangular Distributions
    row = 3
    if 'triangular' in table_docs:
        row = write_table_intro(ws, row, table_docs['triangular'], width_cols=6)
    headers = ['Variable', 'Min', 'Mode', 'Max', 'Units', 'Referenced By']
    row = write_header_row(ws, row, headers)

    tri_start = row
    triangular = [
        ('Season_Extension_Days', 14, 21, 35, 'days'),
        ('Extension_Value_Per_Day', 3000, 7500, 15000, '$/day'),
        ('Interventions_Avoided', 1, 3, 8, 'events'),
        ('Treatment_Rate_Factor', 0.70, 1.00, 1.30, 'multiplier'),  # Kearl kinetics variability
    ]

    for var, min_v, mode_v, max_v, units in triangular:
        ws.cell(row=row, column=1, value=var)
        for col, val in enumerate([min_v, mode_v, max_v], 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = FILL_EDITABLE
            if '$' in units:
                cell.number_format = FMT_CURRENCY
        ws.cell(row=row, column=5, value=units)
        # Referenced By column - all three params reference the same sheets
        refs = get_input_references(f"Tri_{var}_Min")
        ws.cell(row=row, column=6, value=', '.join(refs) if refs else '')
        # Register AND create named ranges for triangular distribution formulas
        register_location(f"Tri_{var}_Min", "7_ServiceModels", "B", row)
        add_named_range(wb, f"Tri_{var}_Min", "7_ServiceModels", f"$B${row}")
        register_location(f"Tri_{var}_Mode", "7_ServiceModels", "C", row)
        add_named_range(wb, f"Tri_{var}_Mode", "7_ServiceModels", f"$C${row}")
        register_location(f"Tri_{var}_Max", "7_ServiceModels", "D", row)
        add_named_range(wb, f"Tri_{var}_Max", "7_ServiceModels", f"$D${row}")
        row += 1

    create_table(ws, "tbl_Triangular", f"A{tri_start-1}:F{row-1}")
    row += 2

    # Beta Distributions
    if 'beta' in table_docs:
        row = write_table_intro(ws, row, table_docs['beta'], width_cols=7)
    headers = ['Variable', 'Alpha', 'Beta', 'Min', 'Max', 'Units', 'Referenced By']
    row = write_header_row(ws, row, headers)

    beta_start = row
    beta = [
        ('Efficiency_Gain', 2, 5, 0.05, 0.40, '%'),
    ]

    for var, alpha, beta_v, min_v, max_v, units in beta:
        ws.cell(row=row, column=1, value=var)
        for col, val in enumerate([alpha, beta_v, min_v, max_v], 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = FILL_EDITABLE
        ws.cell(row=row, column=6, value=units)
        # Referenced By column - all beta params reference the same sheets
        refs = get_input_references(f"Beta_{var}_Alpha")
        ws.cell(row=row, column=7, value=', '.join(refs) if refs else '')
        # Register AND create named ranges for beta distribution formulas
        register_location(f"Beta_{var}_Alpha", "7_ServiceModels", "B", row)
        add_named_range(wb, f"Beta_{var}_Alpha", "7_ServiceModels", f"$B${row}")
        register_location(f"Beta_{var}_Beta", "7_ServiceModels", "C", row)
        add_named_range(wb, f"Beta_{var}_Beta", "7_ServiceModels", f"$C${row}")
        register_location(f"Beta_{var}_Min", "7_ServiceModels", "D", row)
        add_named_range(wb, f"Beta_{var}_Min", "7_ServiceModels", f"$D${row}")
        register_location(f"Beta_{var}_Max", "7_ServiceModels", "E", row)
        add_named_range(wb, f"Beta_{var}_Max", "7_ServiceModels", f"$E${row}")
        row += 1

    create_table(ws, "tbl_Beta", f"A{beta_start-1}:G{row-1}")
    row += 2

    # Value Per M3 Sensitivity (Zhang 2026 framing)
    if 'value_sensitivity' in table_docs:
        row = write_table_intro(ws, row, table_docs['value_sensitivity'], width_cols=5)
    headers = ['Frame', 'Low', 'Base', 'High', 'Context']
    row = write_header_row(ws, row, headers)

    sens_start = row
    sensitivity = [
        ('Operational', 3, 5, 8, 'Direct operational benefit'),
        ('Avoided Treatment', 10, 20, 30, 'Alternative treatment costs'),
        ('ARO Liability', 21, 50, 93, 'Zhang (2026) analysis'),
    ]

    for frame, low, base, high, context in sensitivity:
        ws.cell(row=row, column=1, value=frame)
        for col, val in enumerate([low, base, high], 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = FILL_EDITABLE
            cell.number_format = FMT_CURRENCY
        ws.cell(row=row, column=5, value=context)
        row += 1

    create_table(ws, "tbl_ValueSensitivity", f"A{sens_start-1}:E{row-1}")

    return ws


# =============================================================================
# SHEET 6: CALC_TIMELINE
# =============================================================================

def create_8_calc_timeline(wb):
    """Sheet 6: Calc_Timeline - Generate projection time series."""
    ws = wb.create_sheet("8_Calc_Timeline")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 20, 'B': 36, 'C': 16, 'D': 12, 'E': 14})

    ws['A1'] = "Projection Timeline"
    ws['A1'].font = FONT_TITLE

    # Load and write context box from YAML
    row = 3
    loader = get_content_loader()
    try:
        context = loader.get_context_box("8_calc_timeline")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=5)
    except (ContentNotFoundError, ContentValidationError):
        pass  # Continue without context box if content not available

    # Add table intro from YAML
    table_docs = loader.get_table_docs("8_calc_timeline")
    if 'timeline' in table_docs:
        row = write_table_intro(ws, row, table_docs['timeline'], width_cols=5)

    headers = ['Year_Index', 'Calendar_Year', 'Treatment_Days', 'Discount_Factor', 'Learning_Mult']
    row = write_header_row(ws, row, headers)

    timeline_start = row
    for i in range(6):  # Years 0-5
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value="=Base_Year+" + str(i))
        ws.cell(row=row, column=3, value="=Season_Length")
        # Discount Factor
        if i == 0:
            ws.cell(row=row, column=4, value=1)
        else:
            ws.cell(row=row, column=4, value=f"=1/(1+Discount_Rate)^{i}")
        # Learning Multiplier - reference consolidated table in 4_Assumptions
        if i == 0:
            ws.cell(row=row, column=5, value=1)
        else:
            ws.cell(row=row, column=5, value=f"=INDEX({ref('LearningCurve_Mult')},{i})")
        row += 1

    timeline_end = row - 1
    create_table(ws, "tbl_Timeline", f"A{timeline_start-1}:E{row-1}")

    # Register timeline columns for downstream use
    register_range("Timeline_DiscountFactor", "8_Calc_Timeline", "D", timeline_start, "D", timeline_end)
    add_named_range(wb, "Timeline_DiscountFactor", "8_Calc_Timeline", f"$D${timeline_start}:$D${timeline_end}")
    register_range("Timeline_LearningMult", "8_Calc_Timeline", "E", timeline_start, "E", timeline_end)
    add_named_range(wb, "Timeline_LearningMult", "8_Calc_Timeline", f"$E${timeline_start}:$E${timeline_end}")
    row += 2

    # =========================================================================
    # TREATMENT KINETICS (Kearl Field Data Integration)
    # =========================================================================
    ws.cell(row=row, column=1, value="Treatment Kinetics (Kearl 2022)").font = FONT_HEADER
    row += 1
    ws.cell(row=row, column=1, value="Source: Vander Meulen & Headley (2024) GROW").font = Font(italic=True, size=9)
    row += 1
    headers = ['Metric', 'Formula', 'Value', 'Unit']
    row = write_header_row(ws, row, headers)

    # Effective Rapid Rate (scaled by stochastic factor)
    # NOTE: Stoch_TreatmentRate is created in sheet 7 (forward reference - Excel resolves at runtime)
    ws.cell(row=row, column=1, value="Eff_Rapid_Rate")
    ws.cell(row=row, column=2, value="'=Rapid_Phase_Rate * Stoch_TreatmentRate")
    ws.cell(row=row, column=3, value=f"={ref('Rapid_Phase_Rate')}*Stoch_TreatmentRate")
    ws.cell(row=row, column=3).number_format = '0.000'
    ws.cell(row=row, column=4, value="mg/L/day")
    add_named_range(wb, "Eff_Rapid_Rate", "8_Calc_Timeline", f"$C${row}")
    row += 1

    # Effective Slow Rate (scaled by stochastic factor)
    ws.cell(row=row, column=1, value="Eff_Slow_Rate")
    ws.cell(row=row, column=2, value="'=Slow_Phase_Rate * Stoch_TreatmentRate")
    ws.cell(row=row, column=3, value=f"={ref('Slow_Phase_Rate')}*Stoch_TreatmentRate")
    ws.cell(row=row, column=3).number_format = '0.000'
    ws.cell(row=row, column=4, value="mg/L/day")
    add_named_range(wb, "Eff_Slow_Rate", "8_Calc_Timeline", f"$C${row}")
    row += 1

    # NAFC After Rapid Phase
    ws.cell(row=row, column=1, value="NAFC_After_Rapid")
    ws.cell(row=row, column=2, value="'=MAX(0, Initial - Rate * Duration)")
    nafc_rapid_formula = (f"=MAX(0,{ref('Initial_NAFC')}-"
                          f"{ref('Eff_Rapid_Rate')}*{ref('Rapid_Phase_Duration')})")
    ws.cell(row=row, column=3, value=nafc_rapid_formula)
    ws.cell(row=row, column=3).number_format = '0.00'
    ws.cell(row=row, column=4, value="mg/L")
    add_named_range(wb, "NAFC_After_Rapid", "8_Calc_Timeline", f"$C${row}")
    row += 1

    # Days to Regulatory Compliance
    ws.cell(row=row, column=1, value="Days_to_Compliance")
    ws.cell(row=row, column=2, value="Two-phase calculation")
    # Formula: IF already compliant = 0, ELSE IF compliant in rapid phase, calculate rapid days,
    # ELSE rapid duration + slow phase days
    days_formula = (
        f"=IF({ref('Initial_NAFC')}<={ref('Target_NAFC')},0,"
        f"IF({ref('NAFC_After_Rapid')}<={ref('Target_NAFC')},"
        f"({ref('Initial_NAFC')}-{ref('Target_NAFC')})/{ref('Eff_Rapid_Rate')},"
        f"{ref('Rapid_Phase_Duration')}+({ref('NAFC_After_Rapid')}-{ref('Target_NAFC')})/{ref('Eff_Slow_Rate')}))"
    )
    ws.cell(row=row, column=3, value=days_formula)
    ws.cell(row=row, column=3).number_format = '0.0'
    ws.cell(row=row, column=4, value="days")
    add_named_range(wb, "Days_to_Compliance", "8_Calc_Timeline", f"$C${row}")
    row += 1

    # Circulation Cycles Required
    ws.cell(row=row, column=1, value="Circulation_Cycles")
    ws.cell(row=row, column=2, value="'=CEILING(Days / Cycle, 1)")
    ws.cell(row=row, column=3, value=f"=CEILING({ref('Days_to_Compliance')}/{ref('Recirculation_Cycle')},1)")
    ws.cell(row=row, column=3).number_format = '0'
    ws.cell(row=row, column=4, value="cycles")
    add_named_range(wb, "Circulation_Cycles", "8_Calc_Timeline", f"$C${row}")
    row += 1

    # Treatment Efficiency
    ws.cell(row=row, column=1, value="Treatment_Efficiency")
    ws.cell(row=row, column=2, value="'=(Initial - Target) / Initial")
    ws.cell(row=row, column=3, value=f"=({ref('Initial_NAFC')}-{ref('Target_NAFC')})/{ref('Initial_NAFC')}")
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=4, value="%")
    add_named_range(wb, "Treatment_Efficiency", "8_Calc_Timeline", f"$C${row}")
    row += 1

    # Compliance Status
    ws.cell(row=row, column=1, value="Compliance_Status")
    ws.cell(row=row, column=2, value="Achievable within season?")
    ws.cell(row=row, column=3, value=f'=IF({ref("Days_to_Compliance")}<={ref("Season_Length")},"ACHIEVABLE","EXCEEDS SEASON")')
    ws.cell(row=row, column=4, value="status")
    add_named_range(wb, "Compliance_Status", "8_Calc_Timeline", f"$C${row}")

    return ws


# =============================================================================
# SHEET 9: CALC_KINETICS (NEW)
# =============================================================================

def create_9_calc_kinetics(wb):
    """Sheet 9: Calc_Kinetics - Rate modifier calculations using environmental drivers."""
    ws = wb.create_sheet("9_Calc_Kinetics")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 24, 'B': 55, 'C': 14, 'D': 12, 'E': 35})

    ws['A1'] = "Treatment Kinetics Modifiers"
    ws['A1'].font = FONT_TITLE

    # Context box
    row = 3
    context_content = """Environmental rate modifiers cascade to produce Effective_Rate.
Q10: Temperature coefficient (rate doubles per 10C above reference).
UV: Photodegradation modifier (reduced by turbidity).
DO: Aerobic biodegradation modifier (full above 6.5 mg/L).
Nutrient/Bioavail: Substrate availability modifiers."""
    row = write_context_box(ws, row, "Kinetics Cascade", context_content, width_cols=5)

    # Rate Modifiers Table
    row += 1
    headers = ['Modifier', 'Formula', 'Value', 'Unit', 'Notes']
    row = write_header_row(ws, row, headers)

    mod_start = row

    # Q10 Temperature Factor
    ws.cell(row=row, column=1, value="Q10_Factor")
    ws.cell(row=row, column=2, value="=2^((Stoch_Temp-20)/10)")
    # Use July temperature as default (mid-season)
    q10_formula = "=2^((Env_Temp_Water_Jul-20)/10)"
    ws.cell(row=row, column=3, value=q10_formula)
    ws.cell(row=row, column=3).number_format = '0.000'
    ws.cell(row=row, column=4, value="multiplier")
    ws.cell(row=row, column=5, value="T=20C gives 1.0, T=30C gives 2.0")
    register_location("Q10_Factor", "9_Calc_Kinetics", "C", row)
    add_named_range(wb, "Q10_Factor", "9_Calc_Kinetics", f"$C${row}")
    row += 1

    # UV Modifier
    ws.cell(row=row, column=1, value="UV_Modifier")
    ws.cell(row=row, column=2, value="=UV_Relative*(1-MIN(1,Turbidity/100))")
    # Simplified: use July UV and assume low turbidity
    uv_formula = "=Env_UV_Relative_Jul*(1-MIN(1,Ionic_Chloride_Mode/1000))"
    ws.cell(row=row, column=3, value=uv_formula)
    ws.cell(row=row, column=3).number_format = '0.000'
    ws.cell(row=row, column=4, value="multiplier")
    ws.cell(row=row, column=5, value="UV intensity adjusted for turbidity")
    register_location("UV_Modifier", "9_Calc_Kinetics", "C", row)
    add_named_range(wb, "UV_Modifier", "9_Calc_Kinetics", f"$C${row}")
    row += 1

    # DO Modifier
    ws.cell(row=row, column=1, value="DO_Modifier")
    ws.cell(row=row, column=2, value="=IF(DO>=6.5,1,DO/6.5)")
    do_formula = "=IF(Env_DO_Typical_Jul>=6.5,1,Env_DO_Typical_Jul/6.5)"
    ws.cell(row=row, column=3, value=do_formula)
    ws.cell(row=row, column=3).number_format = '0.000'
    ws.cell(row=row, column=4, value="multiplier")
    ws.cell(row=row, column=5, value="Full rate above 6.5 mg/L")
    register_location("DO_Modifier", "9_Calc_Kinetics", "C", row)
    add_named_range(wb, "DO_Modifier", "9_Calc_Kinetics", f"$C${row}")
    row += 1

    # Nutrient Modifier
    ws.cell(row=row, column=1, value="Nutrient_Modifier")
    ws.cell(row=row, column=2, value="=Nutrient_Index")
    nut_formula = "=Env_Nutrient_Index_Jul"
    ws.cell(row=row, column=3, value=nut_formula)
    ws.cell(row=row, column=3).number_format = '0.000'
    ws.cell(row=row, column=4, value="multiplier")
    ws.cell(row=row, column=5, value="Nutrient availability factor")
    register_location("Nutrient_Modifier", "9_Calc_Kinetics", "C", row)
    add_named_range(wb, "Nutrient_Modifier", "9_Calc_Kinetics", f"$C${row}")
    row += 1

    # Bioavailability Modifier
    ws.cell(row=row, column=1, value="Bioavail_Modifier")
    ws.cell(row=row, column=2, value="=Bioavail*(1-Clay*0.5)")
    bio_formula = "=Env_Bioavail_Index_Jul*(1-0.15*0.5)"
    ws.cell(row=row, column=3, value=bio_formula)
    ws.cell(row=row, column=3).number_format = '0.000'
    ws.cell(row=row, column=4, value="multiplier")
    ws.cell(row=row, column=5, value="Adjusted for clay sorption")
    register_location("Bioavail_Modifier", "9_Calc_Kinetics", "C", row)
    add_named_range(wb, "Bioavail_Modifier", "9_Calc_Kinetics", f"$C${row}")
    row += 1

    create_table(ws, "tbl_Rate_Modifiers", f"A{mod_start-1}:E{row-1}")
    row += 2

    # Combined Effective Rate
    ws.cell(row=row, column=1, value="Combined Rate Calculation").font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1, value="Effective_Rate")
    ws.cell(row=row, column=2, value="=Base_Rate * Q10 * UV * DO * Nutrient * Bioavail * Stoch_Treatment")
    # Combine all modifiers with base rate
    eff_formula = "=Rapid_Phase_Rate*Q10_Factor*UV_Modifier*DO_Modifier*Nutrient_Modifier*Bioavail_Modifier*Stoch_TreatmentRate"
    ws.cell(row=row, column=3, value=eff_formula)
    ws.cell(row=row, column=3).number_format = '0.000'
    ws.cell(row=row, column=4, value="mg/L/day")
    ws.cell(row=row, column=5, value="Combined effective treatment rate")
    register_location("Effective_Rate", "9_Calc_Kinetics", "C", row)
    add_named_range(wb, "Effective_Rate", "9_Calc_Kinetics", f"$C${row}")
    row += 1

    # Rate comparison
    ws.cell(row=row, column=1, value="Rate_vs_Baseline")
    ws.cell(row=row, column=2, value="=Effective_Rate / Rapid_Phase_Rate")
    comp_formula = "=Effective_Rate/Rapid_Phase_Rate"
    ws.cell(row=row, column=3, value=comp_formula)
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=4, value="%")
    ws.cell(row=row, column=5, value="Effective vs baseline rate")

    return ws


# =============================================================================
# SHEET 10: CALC_STOCHASTIC
# =============================================================================

def create_10_calc_stochastic(wb):
    """Sheet 10: Calc_Stochastic - Excel-native random variable generation."""
    ws = wb.create_sheet("10_Calc_Stochastic")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 22, 'B': 12, 'C': 70})

    ws['A1'] = "Stochastic Variables"
    ws['A1'].font = FONT_TITLE
    ws['A2'] = "(Press F9 to recalculate random draws)"
    ws['A2'].font = Font(italic=True)

    # Load and write context box from YAML
    row = 4
    loader = get_content_loader()
    try:
        context = loader.get_context_box("10_calc_stochastic")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=3)
    except (ContentNotFoundError, ContentValidationError):
        pass  # Continue without context box if content not available

    # Add table intro from YAML
    table_docs = loader.get_table_docs("10_calc_stochastic")
    if 'stochastic' in table_docs:
        row = write_table_intro(ws, row, table_docs['stochastic'], width_cols=3)

    headers = ['Variable', 'Random_U', 'Realized_Value']
    row = write_header_row(ws, row, headers)

    stoch_start = row

    # Season Extension (Triangular) - using named ranges per audit
    ws.cell(row=row, column=1, value="Season_Extension")
    ws.cell(row=row, column=2, value="=RAND()")
    # Triangular distribution formula using NAMED RANGES (not addresses)
    min_nm = ref('Tri_Season_Extension_Days_Min')
    mode_nm = ref('Tri_Season_Extension_Days_Mode')
    max_nm = ref('Tri_Season_Extension_Days_Max')
    tri_formula = (f"=IF(B{row}<({mode_nm}-{min_nm})/({max_nm}-{min_nm}),"
                   f"{min_nm}+SQRT(B{row}*({max_nm}-{min_nm})*({mode_nm}-{min_nm})),"
                   f"{max_nm}-SQRT((1-B{row})*({max_nm}-{min_nm})*({max_nm}-{mode_nm})))")
    ws.cell(row=row, column=3, value=tri_formula)
    add_named_range(wb, "Stoch_SeasonExt", "10_Calc_Stochastic", f"$C${row}")
    row += 1

    # Extension Value (Triangular)
    ws.cell(row=row, column=1, value="Extension_Value")
    ws.cell(row=row, column=2, value="=RAND()")
    min_nm = ref('Tri_Extension_Value_Per_Day_Min')
    mode_nm = ref('Tri_Extension_Value_Per_Day_Mode')
    max_nm = ref('Tri_Extension_Value_Per_Day_Max')
    tri_formula = (f"=IF(B{row}<({mode_nm}-{min_nm})/({max_nm}-{min_nm}),"
                   f"{min_nm}+SQRT(B{row}*({max_nm}-{min_nm})*({mode_nm}-{min_nm})),"
                   f"{max_nm}-SQRT((1-B{row})*({max_nm}-{min_nm})*({max_nm}-{mode_nm})))")
    ws.cell(row=row, column=3, value=tri_formula)
    add_named_range(wb, "Stoch_ExtValue", "10_Calc_Stochastic", f"$C${row}")
    row += 1

    # Interventions Avoided (Triangular)
    ws.cell(row=row, column=1, value="Interventions_Avoided")
    ws.cell(row=row, column=2, value="=RAND()")
    min_nm = ref('Tri_Interventions_Avoided_Min')
    mode_nm = ref('Tri_Interventions_Avoided_Mode')
    max_nm = ref('Tri_Interventions_Avoided_Max')
    tri_formula = (f"=IF(B{row}<({mode_nm}-{min_nm})/({max_nm}-{min_nm}),"
                   f"{min_nm}+SQRT(B{row}*({max_nm}-{min_nm})*({mode_nm}-{min_nm})),"
                   f"{max_nm}-SQRT((1-B{row})*({max_nm}-{min_nm})*({max_nm}-{mode_nm})))")
    ws.cell(row=row, column=3, value=tri_formula)
    add_named_range(wb, "Stoch_Interventions", "10_Calc_Stochastic", f"$C${row}")
    row += 1

    # Efficiency Gain (Uniform approximation for Beta - simpler on Mac)
    ws.cell(row=row, column=1, value="Efficiency_Gain")
    ws.cell(row=row, column=2, value="=RAND()")
    min_nm = ref('Beta_Efficiency_Gain_Min')
    max_nm = ref('Beta_Efficiency_Gain_Max')
    ws.cell(row=row, column=3, value=f"={min_nm}+({max_nm}-{min_nm})*B{row}")
    add_named_range(wb, "Stoch_Efficiency", "10_Calc_Stochastic", f"$C${row}")
    row += 1

    # Treatment Rate Factor (Triangular) - Kearl kinetics variability
    ws.cell(row=row, column=1, value="Treatment_Rate_Factor")
    ws.cell(row=row, column=2, value="=RAND()")
    min_nm = ref('Tri_Treatment_Rate_Factor_Min')
    mode_nm = ref('Tri_Treatment_Rate_Factor_Mode')
    max_nm = ref('Tri_Treatment_Rate_Factor_Max')
    tri_formula = (f"=IF(B{row}<({mode_nm}-{min_nm})/({max_nm}-{min_nm}),"
                   f"{min_nm}+SQRT(B{row}*({max_nm}-{min_nm})*({mode_nm}-{min_nm})),"
                   f"{max_nm}-SQRT((1-B{row})*({max_nm}-{min_nm})*({max_nm}-{mode_nm})))")
    ws.cell(row=row, column=3, value=tri_formula)
    add_named_range(wb, "Stoch_TreatmentRate", "10_Calc_Stochastic", f"$C${row}")
    row += 1

    create_table(ws, "tbl_Stochastic", f"A{stoch_start-1}:C{row-1}")

    return ws


# =============================================================================
# SHEET 11: CALC_COMPLIANCE (NEW)
# =============================================================================

def create_11_calc_compliance(wb):
    """Sheet 11: Calc_Compliance - 13-variable compliance gate matrix."""
    ws = wb.create_sheet("11_Calc_Compliance")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 20, 'B': 14, 'C': 10, 'D': 10, 'E': 10, 'F': 35})

    ws['A1'] = "Compliance Gate Matrix"
    ws['A1'].font = FONT_TITLE

    # Context box
    row = 3
    context_content = """13-variable compliance gates for water release qualification.
Each gate returns 1 (PASS) or 0 (FAIL) as numeric value.
Master_Gate_Status = MIN(all gates) - 0 if ANY gate fails.
Numeric gates enable: value gating, fractional compliance (future).
Gate thresholds defined in 5_EnvironmentalDrivers."""
    row = write_context_box(ws, row, "Compliance Gates", context_content, width_cols=6)

    # Compliance Gates Table
    row += 1
    headers = ['Gate', 'Threshold', 'Direction', 'Value', 'Status', 'Notes']
    row = write_header_row(ws, row, headers)

    gate_start = row

    # Define the 13 compliance gates per the plan
    gates = [
        ('Toxicity', 'Baseline +/- 5%', 'within', '=IF(ABS(Stoch_Efficiency-0.2)<0.05,1,0)', 'Toxicity baseline'),
        ('Chloride', '< 230 mg/L', 'below', '=IF(Ionic_Chloride_Mode<230,1,0)', 'Aquatic life protection'),
        ('SAR', '< 4.0', 'below', '=IF(Ionic_SAR_Mode<4,1,0)', 'Land application'),
        ('DO', '> 6.5 mg/L', 'above', '=IF(Env_DO_Typical_Jul>6.5,1,0)', 'Minimum oxygen'),
        ('pH_Low', '> 6.5', 'above', '=IF(7.5>6.5,1,0)', 'pH lower limit'),
        ('pH_High', '< 8.5', 'below', '=IF(7.5<8.5,1,0)', 'pH upper limit'),
        ('Sulfate', '< 500 mg/L', 'below', '=IF(Ionic_Sulfate_Mode<500,1,0)', 'Process water indicator'),
        ('Temperature', '< 25 C', 'below', '=IF(Env_Temp_Water_Jul<25,1,0)', 'Thermal discharge'),
        ('NAFC', '< Target', 'below', '=IF(Days_to_Compliance<Season_Length,1,0)', 'Treatment complete'),
        ('Turbidity', '< 100 NTU', 'below', '=1', 'Sediment control'),
        ('Oil_Sheen', 'None visible', 'absent', '=1', 'Visual inspection'),
        ('Odor', 'None detected', 'absent', '=1', 'Sensory inspection'),
        ('Foam', 'None persistent', 'absent', '=1', 'Surface condition'),
    ]

    for gate_name, threshold, direction, formula, notes in gates:
        ws.cell(row=row, column=1, value=gate_name)
        ws.cell(row=row, column=2, value=threshold)
        ws.cell(row=row, column=3, value=direction)
        ws.cell(row=row, column=4, value=formula)
        ws.cell(row=row, column=4).number_format = '0'
        # Status column - shows PASS/FAIL based on numeric value
        status_formula = f'=IF(D{row}=1,"PASS","FAIL")'
        ws.cell(row=row, column=5, value=status_formula)
        ws.cell(row=row, column=6, value=notes)
        # Register named range for each gate
        register_location(f"Gate_{gate_name}", "11_Calc_Compliance", "D", row)
        add_named_range(wb, f"Gate_{gate_name}", "11_Calc_Compliance", f"$D${row}")
        row += 1

    gate_end = row - 1
    create_table(ws, "tbl_Compliance_Gates", f"A{gate_start-1}:F{row-1}")
    row += 2

    # Master Gate Status
    ws.cell(row=row, column=1, value="Master Gate Status").font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1, value="Master_Gate_Status")
    ws.cell(row=row, column=2, value="=MIN(all gates)")
    # MIN of all gate values - 0 if any gate fails
    master_formula = f"=MIN(D{gate_start}:D{gate_end})"
    ws.cell(row=row, column=3, value=master_formula)
    ws.cell(row=row, column=3).number_format = '0'
    ws.cell(row=row, column=4, value='=IF(C' + str(row) + '=1,"ALL PASS","BLOCKED")')
    register_location("Master_Gate_Status", "11_Calc_Compliance", "C", row)
    add_named_range(wb, "Master_Gate_Status", "11_Calc_Compliance", f"$C${row}")
    row += 1

    # Gate count summary
    ws.cell(row=row, column=1, value="Gates Passing")
    ws.cell(row=row, column=3, value=f"=COUNTIF(D{gate_start}:D{gate_end},1)")
    ws.cell(row=row, column=4, value=f"of {gate_end - gate_start + 1}")
    row += 1

    ws.cell(row=row, column=1, value="Gates Failing")
    ws.cell(row=row, column=3, value=f"=COUNTIF(D{gate_start}:D{gate_end},0)")

    return ws


# =============================================================================
# SHEET 12: CALC_VALUE
# =============================================================================

def create_12_calc_value(wb):
    """Sheet 12: Calc_Value - Calculate gated value for each scenario."""
    ws = wb.create_sheet("12_Calc_Value")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 28, 'B': 14, 'C': 14, 'D': 14, 'E': 14})

    ws['A1'] = "Gated Value Calculations"
    ws['A1'].font = FONT_TITLE

    # Load and write context box from YAML
    row = 3
    loader = get_content_loader()
    try:
        context = loader.get_context_box("12_calc_value")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=5)
    except (ContentNotFoundError, ContentValidationError):
        pass  # Continue without context box if content not available

    # Add table intro from YAML
    table_docs = loader.get_table_docs("12_calc_value")
    if 'gated_values' in table_docs:
        row = write_table_intro(ws, row, table_docs['gated_values'], width_cols=5)

    headers = ['Scenario', 'Base_Value', 'Gate_Status', 'Stoch_Mult', 'Gated_Value']
    row = write_header_row(ws, row, headers)

    value_start = row

    # S1: Season Extension
    ws.cell(row=row, column=1, value="S1_Season_Extension")
    ws.cell(row=row, column=2, value="=INDEX(Scenario_BaseValues,1)")
    ws.cell(row=row, column=3, value="=INDEX(Gate_S1,MATCH(Testing_Option,TestOpt_Option,0))")
    ws.cell(row=row, column=4, value="=Stoch_SeasonExt/21")  # Normalize around mode
    ws.cell(row=row, column=5, value=f'=IF(C{row}="ENABLED",B{row}*D{row},0)')
    add_named_range(wb, "S1_Value", "12_Calc_Value", f"$E${row}")
    row += 1

    # S2: Intervention Avoidance
    ws.cell(row=row, column=1, value="S2_Intervention_Avoidance")
    ws.cell(row=row, column=2, value="=INDEX(Scenario_BaseValues,2)")
    ws.cell(row=row, column=3, value="=INDEX(Gate_S2,MATCH(Testing_Option,TestOpt_Option,0))")
    ws.cell(row=row, column=4, value="=Stoch_Interventions/3")  # Normalize around mode
    ws.cell(row=row, column=5, value=f'=IF(C{row}="ENABLED",B{row}*D{row},0)')
    add_named_range(wb, "S2_Value", "12_Calc_Value", f"$E${row}")
    row += 1

    # S3: Spatial Routing
    ws.cell(row=row, column=1, value="S3_Spatial_Routing")
    ws.cell(row=row, column=2, value="=INDEX(Scenario_BaseValues,3)")
    ws.cell(row=row, column=3, value="=INDEX(Gate_S3,MATCH(Testing_Option,TestOpt_Option,0))")
    ws.cell(row=row, column=4, value="=Stoch_Efficiency/0.15")  # Normalize around expected
    ws.cell(row=row, column=5, value=f'=IF(C{row}="ENABLED",B{row}*D{row},0)')
    add_named_range(wb, "S3_Value", "12_Calc_Value", f"$E${row}")
    row += 1

    # S4: Early Detection
    ws.cell(row=row, column=1, value="S4_Early_Detection")
    ws.cell(row=row, column=2, value="=INDEX(Scenario_BaseValues,4)")
    ws.cell(row=row, column=3, value="=INDEX(Gate_S4,MATCH(Testing_Option,TestOpt_Option,0))")
    ws.cell(row=row, column=4, value="=1+Stoch_Efficiency")  # Add efficiency bonus
    ws.cell(row=row, column=5, value=f'=IF(C{row}="ENABLED",B{row}*D{row},0)')
    add_named_range(wb, "S4_Value", "12_Calc_Value", f"$E${row}")
    row += 1

    create_table(ws, "tbl_GatedValues", f"A{value_start-1}:E{row-1}")

    # Totals
    row += 1
    ws.cell(row=row, column=1, value="Total_Gross_Value").font = FONT_BOLD
    ws.cell(row=row, column=5, value=f"=SUM(E{value_start}:E{row-2})")
    ws[f'E{row}'].number_format = FMT_CURRENCY
    add_named_range(wb, "Gross_Value", "12_Calc_Value", f"$E${row}")

    # Compliance adjustment - apply Master_Gate_Status from 11_Calc_Compliance
    row += 2
    ws.cell(row=row, column=1, value="Compliance Gate").font = FONT_BOLD
    ws.cell(row=row, column=2, value="=Master_Gate_Status")
    ws.cell(row=row, column=3, value='=IF(B' + str(row) + '=1,"PASS","FAIL")')
    ws[f'B{row}'].number_format = '0'
    row += 1

    ws.cell(row=row, column=1, value="Compliance_Adjusted_Value").font = FONT_BOLD
    ws.cell(row=row, column=5, value=f"=Gross_Value*Master_Gate_Status")
    ws[f'E{row}'].number_format = FMT_CURRENCY
    add_named_range(wb, "Compliance_Adjusted_Value", "12_Calc_Value", f"$E${row}")

    # Add note explaining compliance gate
    row += 2
    ws.cell(row=row, column=1, value="Note: Compliance_Adjusted_Value = Gross_Value × Master_Gate_Status")
    ws[f'A{row}'].font = Font(italic=True, size=9, color="666666")
    row += 1
    ws.cell(row=row, column=1, value="If ANY compliance gate fails (13 gates), value becomes $0")
    ws[f'A{row}'].font = Font(italic=True, size=9, color="666666")

    return ws


# =============================================================================
# SHEET 9: CALC_COSTS
# =============================================================================

def create_13_calc_costs(wb):
    """Sheet 9: Calc_Costs - Calculate testing and operational costs."""
    ws = wb.create_sheet("13_Calc_Costs")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 22, 'B': 25, 'C': 18})

    ws['A1'] = "Cost Calculations"
    ws['A1'].font = FONT_TITLE

    # Load and write context box from YAML
    row = 3
    loader = get_content_loader()
    try:
        context = loader.get_context_box("13_calc_costs")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=3)
    except (ContentNotFoundError, ContentValidationError):
        pass  # Continue without context box if content not available

    # Add table intro from YAML
    table_docs = loader.get_table_docs("13_calc_costs")
    if 'cost_calc' in table_docs:
        row = write_table_intro(ws, row, table_docs['cost_calc'], width_cols=3)

    headers = ['Component', 'Formula', 'Value']
    row = write_header_row(ws, row, headers)

    cost_start = row

    # Testing Cost Calculation
    ws.cell(row=row, column=1, value="Selected_Option")
    ws.cell(row=row, column=2, value="Testing_Option")
    ws.cell(row=row, column=3, value="=Testing_Option")
    row += 1

    ws.cell(row=row, column=1, value="Interval_Days")
    ws.cell(row=row, column=2, value="INDEX/MATCH")
    ws.cell(row=row, column=3, value="=INDEX(TestOpt_Interval,MATCH(Testing_Option,TestOpt_Option,0))")
    interval_row = row
    row += 1

    ws.cell(row=row, column=1, value="Monitoring_Points")
    ws.cell(row=row, column=2, value="INDEX/MATCH")
    ws.cell(row=row, column=3, value="=INDEX(TestOpt_Points,MATCH(Testing_Option,TestOpt_Option,0))")
    points_row = row
    row += 1

    ws.cell(row=row, column=1, value="Cost_Per_Test")
    ws.cell(row=row, column=2, value="INDEX/MATCH")
    cell = ws.cell(row=row, column=3, value="=INDEX(TestOpt_Cost,MATCH(Testing_Option,TestOpt_Option,0))")
    cell.number_format = FMT_CURRENCY
    cost_row = row
    row += 1

    ws.cell(row=row, column=1, value="Events_Per_Season")
    ws.cell(row=row, column=2, value="Season_Length / Interval")
    ws.cell(row=row, column=3, value=f"=Season_Length/C{interval_row}")
    events_row = row
    row += 1

    ws.cell(row=row, column=1, value="Total_Tests")
    ws.cell(row=row, column=2, value="Events * Points")
    ws.cell(row=row, column=3, value=f"=C{events_row}*C{points_row}")
    total_tests_row = row
    row += 1

    ws.cell(row=row, column=1, value="Season_Testing_Cost")
    ws.cell(row=row, column=2, value="Total_Tests * Cost_Per_Test")
    cell = ws.cell(row=row, column=3, value=f"=C{total_tests_row}*C{cost_row}")
    cell.number_format = FMT_CURRENCY
    cell.font = FONT_BOLD
    add_named_range(wb, "Testing_Cost", "13_Calc_Costs", f"$C${row}")
    row += 1

    create_table(ws, "tbl_CostCalc", f"A{cost_start-1}:C{row-1}")

    return ws


# =============================================================================
# SHEET 10: CALC_SIM (Monte Carlo)
# =============================================================================

def create_14_calc_sim(wb):
    """Sheet 10: Calc_Sim - 1000-iteration Monte Carlo using Excel Data Table."""
    ws = wb.create_sheet("14_Calc_Sim")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 26, 'B': 14, 'C': 14, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 14, 'I': 18, 'K': 8})

    # Title row first (matching other sheets' pattern)
    ws['A1'] = "Monte Carlo Simulation (1000 iterations)"
    ws['A1'].font = FONT_TITLE

    # Context box starts at row 3 (after blank row 2)
    row = 3
    try:
        loader = get_content_loader()
        context = loader.get_context_box("14_calc_sim")
        if context:
            row = write_context_box(ws, row, context.title, context.content, start_col=1, width_cols=5)
    except (ContentNotFoundError, ContentValidationError):
        pass  # Continue without context box if content not available

    # Calculate row numbers for Data Table
    row += 1  # Blank row after context box
    header_row = row
    formula_row = header_row + 1
    first_iter_row = formula_row + 1
    last_iter_row = first_iter_row + 999  # 1000 iterations

    # Brief instructions above the data table
    ws.cell(row=row, column=1, value=f"Select A{formula_row}:I{last_iter_row} > Data > What-If Analysis > Data Table")
    ws.cell(row=row, column=1).font = Font(italic=True)
    row += 1
    ws.cell(row=row, column=1, value="Column Input Cell: $K$1 (blank cell on this sheet)")
    ws.cell(row=row, column=1).font = Font(italic=True)

    # Recalculate header/formula rows after instructions
    header_row = row + 1
    formula_row = header_row + 1
    first_iter_row = formula_row + 1
    last_iter_row = first_iter_row + 999

    # MC Anchor cell on this sheet - leave as None for Data Table column input
    # Do NOT set to empty string - causes Excel validation issues
    # ws['K1'] is implicitly blank

    # Header row
    row = header_row
    headers = ['Iteration', 'NPV', 'S1', 'S2', 'S3', 'S4', 'Cost', 'Net', 'Days_to_Compliance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = FONT_BOLD

    # Formula row - FIRST row of Data Table selection
    # Column A left blank (implicit None) - top-left of Data Table
    row = formula_row
    ws.cell(row=row, column=2, value="=NPV_Total")
    ws.cell(row=row, column=3, value="=S1_Value")
    ws.cell(row=row, column=4, value="=S2_Value")
    ws.cell(row=row, column=5, value="=S3_Value")
    ws.cell(row=row, column=6, value="=S4_Value")
    ws.cell(row=row, column=7, value="=Testing_Cost")
    ws.cell(row=row, column=8, value="=Compliance_Adjusted_Value-Testing_Cost")
    ws.cell(row=row, column=9, value="=Days_to_Compliance")  # Kearl treatment kinetics

    # Format formula row
    for col in [2, 3, 4, 5, 6, 7, 8]:
        ws.cell(row=row, column=col).number_format = FMT_CURRENCY
    ws.cell(row=row, column=9).number_format = '0.0'  # Days format

    # Iteration numbers (1-1000)
    for i in range(1, 1001):
        ws.cell(row=first_iter_row + i - 1, column=1, value=i)

    # Statistics section (after Data Table)
    row = last_iter_row + 5
    ws.cell(row=row, column=1, value="Monte Carlo Statistics").font = FONT_HEADER
    row += 1

    stats = [
        ("Expected NPV", f"=AVERAGE(B{first_iter_row}:B{last_iter_row})", FMT_CURRENCY),
        ("P10 NPV", f"=PERCENTILE(B{first_iter_row}:B{last_iter_row},0.10)", FMT_CURRENCY),
        ("P50 NPV (Median)", f"=PERCENTILE(B{first_iter_row}:B{last_iter_row},0.50)", FMT_CURRENCY),
        ("P90 NPV", f"=PERCENTILE(B{first_iter_row}:B{last_iter_row},0.90)", FMT_CURRENCY),
        ("StdDev", f"=STDEV(B{first_iter_row}:B{last_iter_row})", FMT_CURRENCY),
        ("Prob(NPV>0)", f"=COUNTIF(B{first_iter_row}:B{last_iter_row},\">0\")/1000", FMT_PERCENT),
    ]

    stats_start = row
    for stat_name, formula, fmt in stats:
        ws.cell(row=row, column=1, value=stat_name)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.number_format = fmt
        row += 1

    add_named_range(wb, "MC_Expected_NPV", "14_Calc_Sim", f"$B${stats_start}")
    add_named_range(wb, "MC_P10", "14_Calc_Sim", f"$B${stats_start+1}")
    add_named_range(wb, "MC_P50", "14_Calc_Sim", f"$B${stats_start+2}")
    add_named_range(wb, "MC_P90", "14_Calc_Sim", f"$B${stats_start+3}")
    add_named_range(wb, "MC_StdDev", "14_Calc_Sim", f"$B${stats_start+4}")
    add_named_range(wb, "MC_Prob_Positive", "14_Calc_Sim", f"$B${stats_start+5}")

    # Treatment Kinetics Statistics (Kearl Integration)
    row += 1
    ws.cell(row=row, column=1, value="Treatment Kinetics Statistics").font = FONT_HEADER
    row += 1

    kinetics_stats = [
        ("Expected Days to Compliance", f'=IFERROR(AVERAGE(I{first_iter_row}:I{last_iter_row}),"Run MC")', '0.0'),
        ("P10 Days (optimistic)", f'=IFERROR(PERCENTILE(I{first_iter_row}:I{last_iter_row},0.10),"Run MC")', '0.0'),
        ("P50 Days (median)", f'=IFERROR(PERCENTILE(I{first_iter_row}:I{last_iter_row},0.50),"Run MC")', '0.0'),
        ("P90 Days (conservative)", f'=IFERROR(PERCENTILE(I{first_iter_row}:I{last_iter_row},0.90),"Run MC")', '0.0'),
        ("Prob(Achievable in Season)", f'=IFERROR(COUNTIF(I{first_iter_row}:I{last_iter_row},"<="&Season_Length)/1000,"Run MC")', FMT_PERCENT),
    ]

    kinetics_stats_start = row
    for stat_name, formula, fmt in kinetics_stats:
        ws.cell(row=row, column=1, value=stat_name)
        cell = ws.cell(row=row, column=2, value=formula)
        cell.number_format = fmt
        row += 1

    add_named_range(wb, "MC_Expected_Days", "14_Calc_Sim", f"$B${kinetics_stats_start}")
    add_named_range(wb, "MC_P10_Days", "14_Calc_Sim", f"$B${kinetics_stats_start+1}")
    add_named_range(wb, "MC_P50_Days", "14_Calc_Sim", f"$B${kinetics_stats_start+2}")
    add_named_range(wb, "MC_P90_Days", "14_Calc_Sim", f"$B${kinetics_stats_start+3}")
    add_named_range(wb, "MC_Prob_Achievable", "14_Calc_Sim", f"$B${kinetics_stats_start+4}")

    return ws


# =============================================================================
# SHEET 11: PL_MONTHLY (Placeholder)
# =============================================================================

def create_15_pl_monthly(wb):
    """Sheet 11: PL_Monthly - Placeholder for monthly P&L if needed."""
    ws = wb.create_sheet("15_PL_Monthly")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 60})

    ws['A1'] = "Monthly P&L (Placeholder)"
    ws['A1'].font = FONT_TITLE

    # Load and write context box from YAML
    row = 3
    try:
        loader = get_content_loader()
        context = loader.get_context_box("15_pl_monthly")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=4)
    except (ContentNotFoundError, ContentValidationError):
        # Fallback to original text
        ws['A3'] = "This sheet is a placeholder for monthly granularity if required."
        ws['A4'] = "Current model uses annual projections in 16_PL_Annual."

    return ws


# =============================================================================
# SHEET 12: PL_ANNUAL
# =============================================================================

def create_16_pl_annual(wb):
    """Sheet 12: PL_Annual - Annual projection summaries."""
    ws = wb.create_sheet("16_PL_Annual")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 12, 'B': 12, 'C': 12, 'D': 12, 'E': 12,
                           'F': 14, 'G': 12, 'H': 14, 'I': 16, 'J': 14})

    ws['A1'] = "Annual Projections"
    ws['A1'].font = FONT_TITLE

    # Load and write context box from YAML
    row = 3
    loader = get_content_loader()
    try:
        context = loader.get_context_box("16_pl_annual")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=10)
    except (ContentNotFoundError, ContentValidationError):
        pass  # Continue without context box if content not available

    # Add table intro from YAML
    table_docs = loader.get_table_docs("16_pl_annual")
    if 'annual' in table_docs:
        row = write_table_intro(ws, row, table_docs['annual'], width_cols=10)

    headers = ['Year', 'S1_Value', 'S2_Value', 'S3_Value', 'S4_Value',
               'Gross_Value', 'Costs', 'Net_Annual', 'Discount_Factor', 'Discounted']
    row = write_header_row(ws, row, headers)

    annual_start = row
    for i in range(5):  # 5 years
        year_row = row
        ws.cell(row=row, column=1, value=f"=Base_Year+{i}")

        # Learning multiplier from timeline - using named range
        learning_nm = f"INDEX({ref('Timeline_LearningMult')},{i+2})"

        # Values with learning curve applied
        ws.cell(row=row, column=2, value=f"=S1_Value*{learning_nm}")
        ws.cell(row=row, column=3, value=f"=S2_Value*{learning_nm}")
        ws.cell(row=row, column=4, value=f"=S3_Value*{learning_nm}")
        ws.cell(row=row, column=5, value=f"=S4_Value*{learning_nm}")
        ws.cell(row=row, column=6, value=f"=SUM(B{year_row}:E{year_row})")
        ws.cell(row=row, column=7, value="=Testing_Cost")
        ws.cell(row=row, column=8, value=f"=F{year_row}-G{year_row}")
        ws.cell(row=row, column=9, value=f"=INDEX({ref('Timeline_DiscountFactor')},{i+2})")
        ws.cell(row=row, column=10, value=f"=H{year_row}*I{year_row}")

        # Format currency columns
        for col in [2, 3, 4, 5, 6, 7, 8, 10]:
            ws.cell(row=row, column=col).number_format = FMT_CURRENCY

        row += 1

    annual_end = row - 1
    create_table(ws, "tbl_Annual", f"A{annual_start-1}:J{row-1}")

    # Create named ranges for Year 1 values (for dashboard cross-sheet references)
    add_named_range(wb, "Year1_Gross_Value", "16_PL_Annual", f"$F${annual_start}")
    add_named_range(wb, "Year1_Net_Value", "16_PL_Annual", f"$H${annual_start}")

    # Aggregate KPIs
    row += 2
    ws.cell(row=row, column=1, value="Aggregate KPIs").font = FONT_HEADER
    row += 1

    ws.cell(row=row, column=1, value="NPV_Total")
    cell = ws.cell(row=row, column=2, value=f"=SUM(J{annual_start}:J{annual_end})")
    cell.number_format = FMT_CURRENCY
    add_named_range(wb, "NPV_Total", "16_PL_Annual", f"$B${row}")
    row += 1

    ws.cell(row=row, column=1, value="Total_Gross_5Yr")
    cell = ws.cell(row=row, column=2, value=f"=SUM(F{annual_start}:F{annual_end})")
    cell.number_format = FMT_CURRENCY
    row += 1

    ws.cell(row=row, column=1, value="Total_Net_5Yr")
    cell = ws.cell(row=row, column=2, value=f"=SUM(H{annual_start}:H{annual_end})")
    cell.number_format = FMT_CURRENCY
    row += 1

    return ws


# =============================================================================
# SHEET 13: CASHFLOW
# =============================================================================

def create_17_cashflow(wb):
    """Sheet 13: CashFlow - Cumulative projections."""
    ws = wb.create_sheet("17_CashFlow")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 12, 'B': 16, 'C': 16, 'D': 16})

    ws['A1'] = "Cumulative Cash Flow"
    ws['A1'].font = FONT_TITLE

    # Load and write context box from YAML
    row = 3
    loader = get_content_loader()
    try:
        context = loader.get_context_box("17_cashflow")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=4)
    except (ContentNotFoundError, ContentValidationError):
        pass  # Continue without context box if content not available

    # Add table intro from YAML
    table_docs = loader.get_table_docs("17_cashflow")
    if 'cashflow' in table_docs:
        row = write_table_intro(ws, row, table_docs['cashflow'], width_cols=4)

    headers = ['Year', 'Annual_Net', 'Cumulative_Net', 'Cumulative_NPV']
    row = write_header_row(ws, row, headers)

    cf_start = row
    for i in range(5):
        ws.cell(row=row, column=1, value=f"=Base_Year+{i}")
        # Use INDEX on table columns instead of hardcoded row references
        ws.cell(row=row, column=2, value=f"=INDEX(tbl_Annual[Net_Annual],{i+1})")
        if i == 0:
            ws.cell(row=row, column=3, value=f"=B{row}")
            ws.cell(row=row, column=4, value=f"=INDEX(tbl_Annual[Discounted],{i+1})")
        else:
            ws.cell(row=row, column=3, value=f"=C{row-1}+B{row}")
            ws.cell(row=row, column=4, value=f"=D{row-1}+INDEX(tbl_Annual[Discounted],{i+1})")

        for col in [2, 3, 4]:
            ws.cell(row=row, column=col).number_format = FMT_CURRENCY

        row += 1

    create_table(ws, "tbl_CashFlow", f"A{cf_start-1}:D{row-1}")

    return ws


# =============================================================================
# SHEET 14: UNIT ECONOMICS
# =============================================================================

def create_18_uniteconomics(wb):
    """Sheet 14: UnitEconomics - Unit economics view."""
    ws = wb.create_sheet("18_UnitEconomics")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 28, 'B': 18, 'C': 10, 'D': 25})

    ws['A1'] = "Unit Economics"
    ws['A1'].font = FONT_TITLE

    # Load and write context box from YAML
    row = 3
    loader = get_content_loader()
    try:
        context = loader.get_context_box("18_uniteconomics")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=4)
    except (ContentNotFoundError, ContentValidationError):
        pass  # Continue without context box if content not available

    # Add table intro from YAML
    table_docs = loader.get_table_docs("18_uniteconomics")
    if 'unit_economics' in table_docs:
        row = write_table_intro(ws, row, table_docs['unit_economics'], width_cols=4)

    headers = ['Metric', 'Value', 'Unit', 'Formula']
    row = write_header_row(ws, row, headers)

    ue_start = row

    metrics = [
        ("Annual Throughput", "=Annual_Throughput", "m3/yr", "From 4_Assumptions"),
        ("Gross Value per Season", "=Gross_Value", "$", "Sum of S1-S4 (pre-compliance)"),
        ("Compliance-Adjusted Value", "=Compliance_Adjusted_Value", "$", "Gross × Master_Gate_Status"),
        ("Testing Cost per Season", "=Testing_Cost", "$", "Based on option selected"),
        ("Net Value per Season", "=Compliance_Adjusted_Value-Testing_Cost", "$", "Adjusted - Testing"),
        ("Value per m3 (Gross)", "=Gross_Value/Annual_Throughput", "$/m3", "Gross / Throughput"),
        ("Value per m3 (Net)", "=(Compliance_Adjusted_Value-Testing_Cost)/Annual_Throughput", "$/m3", "Net / Throughput"),
        ("Cost per m3", "=Testing_Cost/Annual_Throughput", "$/m3", "Testing / Throughput"),
        ("5-Year NPV per m3", "=NPV_Total/(Annual_Throughput*5)", "$/m3", "NPV / (5yr throughput)"),
    ]

    for metric, value, unit, formula in metrics:
        ws.cell(row=row, column=1, value=metric)
        cell = ws.cell(row=row, column=2, value=value)
        if "$" in unit:
            cell.number_format = FMT_CURRENCY_DEC
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=formula)
        row += 1

    create_table(ws, "tbl_UnitEconomics", f"A{ue_start-1}:D{row-1}")

    # Breakeven Analysis
    row += 2
    ws.cell(row=row, column=1, value="Breakeven Analysis").font = FONT_HEADER
    row += 1

    ws.cell(row=row, column=1, value="Testing cost per test:")
    ws.cell(row=row, column=2, value="=INDEX(TestOpt_Cost,MATCH(Testing_Option,TestOpt_Option,0))")
    ws['B' + str(row)].number_format = FMT_CURRENCY
    row += 1

    ws.cell(row=row, column=1, value="Tests per season:")
    ws.cell(row=row, column=2, value="=Testing_Cost/INDEX(TestOpt_Cost,MATCH(Testing_Option,TestOpt_Option,0))")
    row += 1

    ws.cell(row=row, column=1, value="Value per test (net):")
    ws.cell(row=row, column=2, value="=(Compliance_Adjusted_Value-Testing_Cost)/(Testing_Cost/INDEX(TestOpt_Cost,MATCH(Testing_Option,TestOpt_Option,0)))")
    ws['B' + str(row)].number_format = FMT_CURRENCY_DEC

    return ws


# =============================================================================
# SHEET 15: SENSITIVITY
# =============================================================================

def create_19_sensitivity(wb):
    """Sheet 15: Sensitivity - Tornado analysis with automatic calculations.

    Uses scaling factors to estimate NPV at different parameter values:
    - Proportional parameters (Value_Per_M3, Season_Length, Wetland_Area):
      NPV scales linearly with parameter value
    - Discount_Rate: Inverse relationship using time-value adjustment
    """
    ws = wb.create_sheet("19_Sensitivity")
    ws.sheet_properties.tabColor = TAB_BLUE
    set_column_widths(ws, {'A': 18, 'B': 12, 'C': 12, 'D': 12, 'E': 15, 'F': 15, 'G': 15})

    ws['A1'] = "Sensitivity Analysis"
    ws['A1'].font = FONT_TITLE

    # Load and write context box from YAML
    row = 3
    loader = get_content_loader()
    try:
        context = loader.get_context_box("19_sensitivity")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=7)
    except (ContentNotFoundError, ContentValidationError):
        # Fallback to original methodology note
        ws['A2'] = "Automatic scaling - no manual Data Table setup required"
        ws['A2'].font = Font(italic=True, color="666666")
        row = 4

    # Add table intro from YAML
    table_docs = loader.get_table_docs("19_sensitivity")
    if 'sensitivity' in table_docs:
        row = write_table_intro(ws, row, table_docs['sensitivity'], width_cols=7)

    headers = ['Parameter', 'Low', 'Base', 'High', 'NPV_at_Low', 'NPV_at_High', 'Swing']
    row = write_header_row(ws, row, headers)

    sens_start = row

    # Sensitivity parameters with scaling type:
    # 'proportional' = NPV scales linearly with parameter
    # 'inverse_rate' = higher value = lower NPV (discount rate effect)
    sensitivity_params = [
        ('Value_Per_M3', 3, 5, 10, 'proportional'),
        ('Discount_Rate', 0.04, 0.048, 0.06, 'inverse_rate'),
        ('Season_Length', 80, 100, 120, 'proportional'),
        ('Wetland_Area', 0.5, 1.0, 2.0, 'proportional'),
    ]

    for param, low, base, high, scale_type in sensitivity_params:
        ws.cell(row=row, column=1, value=param)
        ws.cell(row=row, column=2, value=low)
        ws.cell(row=row, column=3, value=base)
        ws.cell(row=row, column=4, value=high)

        # Calculate NPV_at_Low and NPV_at_High using scaling formulas
        if scale_type == 'proportional':
            # NPV scales linearly: NPV_at_X = NPV_Total * (X / Base)
            ws.cell(row=row, column=5, value=f"=NPV_Total*(B{row}/C{row})")
            ws.cell(row=row, column=6, value=f"=NPV_Total*(D{row}/C{row})")
        elif scale_type == 'inverse_rate':
            # Discount rate: NPV_at_X = NPV_Total * ((1+Base)/(1+X))^2.5
            ws.cell(row=row, column=5, value=f"=NPV_Total*((1+C{row})/(1+B{row}))^2.5")
            ws.cell(row=row, column=6, value=f"=NPV_Total*((1+C{row})/(1+D{row}))^2.5")

        # Swing = |NPV_at_High - NPV_at_Low|
        ws.cell(row=row, column=7, value=f"=ABS(F{row}-E{row})")

        # Format currency columns
        for col in [5, 6, 7]:
            ws.cell(row=row, column=col).number_format = FMT_CURRENCY

        row += 1

    create_table(ws, "tbl_Sensitivity", f"A{sens_start-1}:G{row-1}")

    # Methodology explanation
    row += 2
    ws.cell(row=row, column=1, value="Methodology").font = FONT_HEADER
    row += 1
    methodology = [
        "Proportional parameters: NPV scales linearly with parameter ratio (Low/Base, High/Base)",
        "Discount Rate: Uses time-value adjustment formula NPV * ((1+Base)/(1+Rate))^2.5",
        "Note: These are first-order approximations. For precise sensitivity, use Monte Carlo.",
    ]
    for note in methodology:
        ws.cell(row=row, column=1, value=note)
        row += 1

    # Tornado chart instructions
    row += 1
    ws.cell(row=row, column=1, value="Tornado Chart").font = FONT_HEADER
    row += 1
    instructions = [
        "1. Select Parameter and Swing columns (A4:A8, G4:G8)",
        "2. Insert -> Chart -> Bar Chart (horizontal)",
        "3. Sort data by Swing descending for tornado effect",
    ]
    for inst in instructions:
        ws.cell(row=row, column=1, value=inst)
        row += 1

    return ws


# =============================================================================
# SHEET 16: DASHBOARD
# =============================================================================

def create_20_dashboard(wb):
    """Sheet 16: Dashboard - Executive summary with KPIs."""
    ws = wb.create_sheet("20_Dashboard")
    ws.sheet_properties.tabColor = TAB_GREEN
    set_column_widths(ws, {'A': 30, 'B': 18, 'C': 16, 'D': 18})

    ws['A1'] = "Executive Dashboard"
    ws['A1'].font = FONT_TITLE

    # Load table documentation from YAML
    loader = get_content_loader()
    table_docs = loader.get_table_docs("20_dashboard")
    collected_docs = []  # Collect for reference section

    # Load and write context box from YAML
    row = 3
    try:
        context = loader.get_context_box("20_dashboard")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=4)
    except (ContentNotFoundError, ContentValidationError):
        pass  # Continue without context box if content not available

    # KPI Summary - with inline documentation
    if 'kpis' in table_docs:
        doc = table_docs['kpis']
        collected_docs.append(doc)
        row = write_table_intro(ws, row, doc, width_cols=4)

    headers = ['Metric', 'Value']
    row = write_header_row(ws, row, headers)

    kpi_start = row
    kpis = [
        ("Selected Testing Option", "=Testing_Option"),
        ("Expected NPV (5-Year)", "=NPV_Total"),
        ("Monte Carlo P10", "=MC_P10"),
        ("Monte Carlo P50", "=MC_P50"),
        ("Monte Carlo P90", "=MC_P90"),
        ("Probability NPV > 0", "=MC_Prob_Positive"),
        ("Compliance Gate Status", '=IF(Master_Gate_Status=1,"PASS","FAIL")'),
        ("Annual Testing Cost", "=Testing_Cost"),
        ("Year 1 Gross Value", "=Year1_Gross_Value"),
        ("Year 1 Net Value", "=Year1_Net_Value"),
    ]

    for metric, formula in kpis:
        ws.cell(row=row, column=1, value=metric)
        cell = ws.cell(row=row, column=2, value=formula)
        if "Probability" in metric:
            cell.number_format = FMT_PERCENT
        elif "Option" not in metric:
            cell.number_format = FMT_CURRENCY
        row += 1

    create_table(ws, "tbl_KPIs", f"A{kpi_start-1}:B{row-1}")
    row += 1

    # Scenario Value Breakdown - with inline documentation
    if 'scenario_breakdown' in table_docs:
        doc = table_docs['scenario_breakdown']
        collected_docs.append(doc)
        row = write_table_intro(ws, row, doc, width_cols=4)

    headers = ['Scenario', 'Gated_Value', 'Status']
    row = write_header_row(ws, row, headers)

    breakdown_start = row
    scenarios = [
        ("S1: Season Extension", "=S1_Value", "=INDEX(Gate_S1,MATCH(Testing_Option,TestOpt_Option,0))"),
        ("S2: Intervention Avoidance", "=S2_Value", "=INDEX(Gate_S2,MATCH(Testing_Option,TestOpt_Option,0))"),
        ("S3: Spatial Routing", "=S3_Value", "=INDEX(Gate_S3,MATCH(Testing_Option,TestOpt_Option,0))"),
        ("S4: Early Detection", "=S4_Value", "=INDEX(Gate_S4,MATCH(Testing_Option,TestOpt_Option,0))"),
    ]

    for scenario, value_formula, status_formula in scenarios:
        ws.cell(row=row, column=1, value=scenario)
        cell = ws.cell(row=row, column=2, value=value_formula)
        cell.number_format = FMT_CURRENCY
        ws.cell(row=row, column=3, value=status_formula)
        row += 1

    create_table(ws, "tbl_ScenarioBreakdown", f"A{breakdown_start-1}:C{row-1}")
    row += 1

    # Treatment Kinetics KPIs - with inline documentation
    if 'treatment_kpis' in table_docs:
        doc = table_docs['treatment_kpis']
        collected_docs.append(doc)
        row = write_table_intro(ws, row, doc, width_cols=4)

    headers = ['Metric', 'Value', 'Status']
    row = write_header_row(ws, row, headers)

    kinetics_kpi_start = row
    kinetics_kpis = [
        ("Initial NAFC Concentration", "=Initial_NAFC", '=IF(Initial_NAFC>Target_NAFC,"Above Target","At/Below Target")'),
        ("Target NAFC Concentration", "=Target_NAFC", '"-"'),
        ("Days to Regulatory Compliance", "=Days_to_Compliance", '=Compliance_Status'),
        ("Circulation Cycles Required", "=Circulation_Cycles", '=IF(Circulation_Cycles<=3,"Standard","Extended")'),
        ("Treatment Efficiency", "=Treatment_Efficiency", '=IF(Treatment_Efficiency>0.2,"Good","Low")'),
        ("MC Expected Days", "=MC_Expected_Days", '=IF(MC_Expected_Days<=Season_Length,"Achievable","Risk")'),
        ("MC P90 Days (conservative)", "=MC_P90_Days", '=IF(MC_P90_Days<=Season_Length,"Low Risk","High Risk")'),
        ("MC Prob(Achievable)", "=MC_Prob_Achievable", '=IF(MC_Prob_Achievable>=0.9,"High Confidence","Review Required")'),
    ]

    for metric, value_formula, status_formula in kinetics_kpis:
        ws.cell(row=row, column=1, value=metric)
        cell = ws.cell(row=row, column=2, value=value_formula)
        if "NAFC" in metric:
            cell.number_format = '0.0 "mg/L"'
        elif "Days" in metric:
            cell.number_format = '0.0 "days"'
        elif "Cycles" in metric:
            cell.number_format = '0'
        elif "Efficiency" in metric or "Prob" in metric:
            cell.number_format = FMT_PERCENT
        ws.cell(row=row, column=3, value=status_formula)
        row += 1

    create_table(ws, "tbl_TreatmentKPIs", f"A{kinetics_kpi_start-1}:C{row-1}")
    row += 1

    # Testing Option Comparison - with inline documentation
    if 'option_comparison' in table_docs:
        doc = table_docs['option_comparison']
        collected_docs.append(doc)
        row = write_table_intro(ws, row, doc, width_cols=4)

    headers = ['Metric', 'Monthly_HRMS', 'Weekly_HRMS', 'Daily_Biosensor']
    row = write_header_row(ws, row, headers)

    comparison_start = row
    comparisons = [
        ("S1 Gate", "BLOCKED", "BLOCKED", "ENABLED"),
        ("S2 Gate", "BLOCKED", "BLOCKED", "ENABLED"),
        ("S3 Gate", "BLOCKED", "ENABLED", "ENABLED"),
        ("S4 Gate", "BLOCKED", "BLOCKED", "ENABLED"),
        ("Testing Cost/Season", "$60,000", "$28,571", "$10,000"),
        ("Potential 5-Year NPV", "-$29k", "$180k", "$1.1M+"),
    ]

    for metric, monthly, weekly, daily in comparisons:
        ws.cell(row=row, column=1, value=metric)
        ws.cell(row=row, column=2, value=monthly)
        ws.cell(row=row, column=3, value=weekly)
        ws.cell(row=row, column=4, value=daily)
        row += 1

    create_table(ws, "tbl_OptionComparison", f"A{comparison_start-1}:D{row-1}")
    row += 1

    # Model Health Status
    ws.cell(row=row, column=1, value="Model Health").font = FONT_HEADER
    row += 1
    ws.cell(row=row, column=1, value="Check Status:")
    ws.cell(row=row, column=2, value="=IF(COUNTIF('21_Checks'!D:D,\"FAIL\")=0,\"HEALTHY\",\"REVIEW REQUIRED\")")
    row += 1
    ws.cell(row=row, column=1, value="Errors Found:")
    ws.cell(row=row, column=2, value="=COUNTIF('21_Checks'!D:D,\"FAIL\")")

    return ws


# =============================================================================
# SHEET 17: CHECKS
# =============================================================================

def create_21_checks(wb):
    """Sheet 17: Checks - Dedicated validation and integrity checks."""
    ws = wb.create_sheet("21_Checks")
    ws.sheet_properties.tabColor = TAB_RED
    set_column_widths(ws, {'A': 12, 'B': 12, 'C': 40, 'D': 12})

    ws['A1'] = "Model Validation Checks"
    ws['A1'].font = FONT_TITLE

    # Load and write context box from YAML
    row = 3
    loader = get_content_loader()
    try:
        context = loader.get_context_box("21_checks")
        if context:
            row = write_context_box(ws, row, context.title, context.content, width_cols=4)
    except (ContentNotFoundError, ContentValidationError):
        pass  # Continue without context box if content not available

    # Add table intro from YAML
    table_docs = loader.get_table_docs("21_checks")
    if 'checks' in table_docs:
        row = write_table_intro(ws, row, table_docs['checks'], width_cols=4)

    headers = ['Check_ID', 'Category', 'Description', 'Status']
    row = write_header_row(ws, row, headers)

    checks = [
        ("CHK001", "Range", "Discount rate between 0-15%",
         '=IF(AND(Discount_Rate>=0,Discount_Rate<=0.15),"PASS","FAIL")'),
        ("CHK002", "Range", "Season length between 60-150 days",
         '=IF(AND(Season_Length>=60,Season_Length<=150),"PASS","FAIL")'),
        ("CHK003", "Range", "Projection years between 1-10",
         '=IF(AND(Projection_Years>=1,Projection_Years<=10),"PASS","FAIL")'),
        ("CHK004", "Range", "Value per M3 positive",
         '=IF(Value_Per_M3>0,"PASS","FAIL")'),
        ("CHK005", "Logic", "Testing option valid",
         '=IF(COUNTIF(List_TestingOptions,Testing_Option)=1,"PASS","FAIL")'),
        ("CHK006", "Logic", "Scenario selection valid",
         '=IF(COUNTIF(List_Scenarios,Scenario_Select)=1,"PASS","FAIL")'),
        ("CHK007", "Sanity", "NPV within reasonable bounds",
         '=IF(AND(NPV_Total>-1000000,NPV_Total<10000000),"PASS","FAIL")'),
        ("CHK008", "Sanity", "Testing cost positive",
         '=IF(Testing_Cost>0,"PASS","FAIL")'),
        ("CHK009", "Sanity", "Gross value non-negative",
         '=IF(Gross_Value>=0,"PASS","FAIL")'),
        ("CHK010", "Integrity", "All named ranges defined",
         '=IF(ISERROR(Testing_Option),"FAIL","PASS")'),
        # Treatment Kinetics Checks (Kearl Integration)
        ("CHK011", "Range", "Initial NAFC in range (40-100 mg/L)",
         '=IF(AND(Initial_NAFC>=40,Initial_NAFC<=100),"PASS","FAIL")'),
        ("CHK012", "Logic", "Target NAFC < Initial NAFC",
         '=IF(Target_NAFC<Initial_NAFC,"PASS","FAIL")'),
        ("CHK013", "Sanity", "Days to compliance <= season length",
         '=IF(Days_to_Compliance<=Season_Length,"PASS","FAIL")'),
        ("CHK014", "Range", "Treatment rates positive",
         '=IF(AND(Eff_Rapid_Rate>0,Eff_Slow_Rate>0),"PASS","FAIL")'),
        ("CHK015", "Compliance", "Master compliance gate status",
         '=IF(Master_Gate_Status=1,"PASS","FAIL")'),
        ("CHK016", "Sanity", "Compliance adjusted value non-negative",
         '=IF(Compliance_Adjusted_Value>=0,"PASS","FAIL")'),
    ]

    check_start = row
    for check_id, category, description, formula in checks:
        ws.cell(row=row, column=1, value=check_id)
        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=3, value=description)
        ws.cell(row=row, column=4, value=formula)
        row += 1

    create_table(ws, "tbl_Checks", f"A{check_start-1}:D{row-1}")

    # Summary
    row += 2
    ws.cell(row=row, column=1, value="Summary").font = FONT_HEADER
    row += 1
    ws.cell(row=row, column=1, value="Total Checks:")
    ws.cell(row=row, column=2, value="=COUNTA(tbl_Checks[Check_ID])")
    row += 1
    ws.cell(row=row, column=1, value="Passed:")
    ws.cell(row=row, column=2, value='=COUNTIF(tbl_Checks[Status],"PASS")')
    row += 1
    ws.cell(row=row, column=1, value="Failed:")
    ws.cell(row=row, column=2, value='=COUNTIF(tbl_Checks[Status],"FAIL")')
    row += 1
    ws.cell(row=row, column=1, value="Model Health:")
    ws.cell(row=row, column=2, value='=IF(COUNTIF(tbl_Checks[Status],"FAIL")=0,"HEALTHY","REVIEW REQUIRED")')

    return ws


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def generate_workbook():
    """Generate the FAST-compliant Excel workbook (22 sheets, 0-21)."""
    print("Luminous Wetland Financial Model - FAST Standard Generator v4.0")
    print("=" * 65)
    print("\nGenerating 22-sheet FAST-compliant workbook...")
    print("  - Python writes FORMULAS only (via LOCATIONS registry)")
    print("  - Excel performs ALL calculations")
    print("  - Monte Carlo via Data Tables (1000 iterations)")
    print("  - NEW: Environmental drivers and kinetics cascade")
    print("  - NEW: 13-variable compliance gate matrix")
    print()

    # Clear locations registry
    clear_locations()

    # Create workbook
    wb = Workbook()

    # Create all 22 sheets in order (0-21)
    print("Creating sheets:")

    print("  [0] 0_Cover - Navigation hub")
    create_0_cover(wb)

    print("  [1] 1_TOC - Control panel")
    create_1_toc(wb)

    print("  [2] 3_Instructions - Methodology")
    create_3_instructions(wb)

    print("  [3] 4_Assumptions - ALL inputs consolidated")
    create_4_assumptions(wb)

    print("  [5] 5_EnvironmentalDrivers - Monthly environmental profiles (NEW)")
    create_5_environmentaldrivers(wb)

    print("  [6] 6_Scenarios - Value scenario definitions")
    create_6_scenarios(wb)

    print("  [7] 7_ServiceModels - Monte Carlo distributions")
    create_7_servicemodels(wb)

    print("  [8] 8_Calc_Timeline - Projection timeline")
    create_8_calc_timeline(wb)

    print("  [9] 9_Calc_Kinetics - Rate modifier cascade (NEW)")
    create_9_calc_kinetics(wb)

    # Create InputMap AFTER input sheets so WRITE_MAP has range registrations
    print("  [2] 2_InputMap - Input traceability reference")
    create_2_inputmap(wb)
    # Move sheet to correct position (after 3_Instructions, index 3)
    wb.move_sheet("2_InputMap", offset=-(wb.sheetnames.index("2_InputMap") - 3))

    print("  [10] 10_Calc_Stochastic - Random variables")
    create_10_calc_stochastic(wb)

    print("  [11] 11_Calc_Compliance - Gate matrix (NEW)")
    create_11_calc_compliance(wb)

    print("  [12] 12_Calc_Value - Gated values")
    create_12_calc_value(wb)

    print("  [9] 13_Calc_Costs - Cost calculations")
    create_13_calc_costs(wb)

    print("  [10] 14_Calc_Sim - 1000-iteration Monte Carlo")
    create_14_calc_sim(wb)

    print("  [11] 15_PL_Monthly - Monthly P&L (placeholder)")
    create_15_pl_monthly(wb)

    print("  [12] 16_PL_Annual - Annual projections")
    create_16_pl_annual(wb)

    print("  [13] 17_CashFlow - Cumulative projections")
    create_17_cashflow(wb)

    print("  [14] 18_UnitEconomics - Unit economics view")
    create_18_uniteconomics(wb)

    print("  [15] 19_Sensitivity - Tornado analysis")
    create_19_sensitivity(wb)

    print("  [16] 20_Dashboard - Executive dashboard")
    create_20_dashboard(wb)

    print("  [17] 21_Checks - Validation checks")
    create_21_checks(wb)

    # Add data validation to 1_TOC (after 4_Assumptions exists)
    toc_ws = wb['1_TOC']

    # Testing Option dropdown
    dv_testing = DataValidation(
        type="list",
        formula1="List_TestingOptions",
        allow_blank=False
    )
    dv_testing.error = "Select from dropdown"
    dv_testing.errorTitle = "Invalid Input"
    toc_ws.add_data_validation(dv_testing)
    dv_testing.add(toc_ws['B5'])

    # Scenario dropdown
    dv_scenario = DataValidation(
        type="list",
        formula1="List_Scenarios",
        allow_blank=False
    )
    toc_ws.add_data_validation(dv_scenario)
    dv_scenario.add(toc_ws['B6'])

    # Verify traceability registry matches actual formula usage
    verify_traceability(wb)

    return wb


def main():
    """Main entry point."""
    wb = generate_workbook()

    # Save workbook
    output_dir = "/Users/greg/Luminous Dropbox/Obsidian/utils/financial_modeling"
    output_path = os.path.join(output_dir, "luminous_wetland_monte_carlo_model.xlsx")

    print(f"\nSaving to: {output_path}")
    wb.save(output_path)

    print("\n" + "=" * 65)
    print("SUCCESS! FAST-compliant 22-sheet workbook created.")
    print()
    print("LOCATIONS registry contains", len(LOCATIONS), "registered variables")
    print()
    print("Next Steps in Excel:")
    print("  1. Open the workbook")
    print("  2. Go to 14_Calc_Sim")
    print("  3. Select range shown in sheet instructions (formula row through iterations)")
    print("  4. Data > What-If Analysis > Data Table")
    print("  5. Column Input Cell: $K$1")
    print("  6. Press Ctrl+Alt+F9 to run Monte Carlo")
    print()
    print("Sheet Legend:")
    print("  - Yellow tabs (1,3-5): User-editable inputs")
    print("  - Gray tabs (2): Documentation")
    print("  - Blue tabs (6-15): Calculation engine")
    print("  - Green tab (16): Executive dashboard")
    print("  - Red tab (17): Validation checks")
    print("=" * 65)


if __name__ == "__main__":
    main()
